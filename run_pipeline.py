"""Run the Research + Analyst Pipeline and write output to Excel.

Usage:
    python run_pipeline.py                           # full pipeline
    python run_pipeline.py data output               # custom dirs (backward compatible)
    python run_pipeline.py --from 5                  # resume from Agent 05
    python run_pipeline.py --from 10 --historical    # resume from Agent 10, include historical
    python run_pipeline.py --data mydata --output results --from 3
"""

import argparse
import sys
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Type

sys.path.insert(0, str(Path(__file__).parent / "src"))

from dotenv import load_dotenv
load_dotenv()

import pandas as pd
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel

from ibd_agents.agents.research_agent import run_research_pipeline
from ibd_agents.agents.analyst_agent import run_analyst_pipeline
from ibd_agents.agents.rotation_detector import run_rotation_pipeline
from ibd_agents.agents.sector_strategist import run_strategist_pipeline
from ibd_agents.agents.portfolio_manager import run_portfolio_pipeline
from ibd_agents.agents.risk_officer import run_risk_pipeline
from ibd_agents.agents.returns_projector import run_returns_projector_pipeline
from ibd_agents.agents.portfolio_reconciler import run_reconciler_pipeline
from ibd_agents.agents.educator_agent import run_educator_pipeline
from ibd_agents.agents.executive_synthesizer import run_synthesizer_pipeline
from ibd_agents.agents.value_investor_agent import run_value_investor_pipeline
from ibd_agents.agents.pattern_agent import run_pattern_pipeline
from ibd_agents.agents.historical_analyst import run_historical_pipeline
from ibd_agents.agents.target_return_constructor import run_target_return_pipeline
from ibd_agents.agents.exit_strategist import run_exit_strategist_pipeline
from ibd_agents.tools.portfolio_reader import read_brokerage_pdfs
from ibd_agents.agents.regime_detector import run_regime_detector_pipeline
from ibd_agents.agents.earnings_risk_analyst import run_earnings_risk_pipeline

# Schema imports for snapshot deserialization
from ibd_agents.schemas.research_output import ResearchOutput
from ibd_agents.schemas.analyst_output import AnalystOutput
from ibd_agents.schemas.rotation_output import RotationDetectionOutput
from ibd_agents.schemas.strategy_output import SectorStrategyOutput
from ibd_agents.schemas.portfolio_output import PortfolioOutput
from ibd_agents.schemas.risk_output import RiskAssessment
from ibd_agents.schemas.returns_projection_output import ReturnsProjectionOutput
from ibd_agents.schemas.reconciliation_output import ReconciliationOutput
from ibd_agents.schemas.educator_output import EducatorOutput
from ibd_agents.schemas.synthesis_output import SynthesisOutput
from ibd_agents.schemas.value_investor_output import ValueInvestorOutput
from ibd_agents.schemas.pattern_output import PortfolioPatternOutput
from ibd_agents.schemas.historical_output import HistoricalAnalysisOutput
from ibd_agents.schemas.target_return_output import TargetReturnOutput
from ibd_agents.schemas.exit_strategy_output import ExitStrategyOutput
from ibd_agents.schemas.regime_detector_output import RegimeDetectorOutput
from ibd_agents.schemas.earnings_risk_output import EarningsRiskOutput
from ibd_agents.tools.token_tracker import tracker as token_tracker


# ---------------------------------------------------------------------------
# Agent Registry — maps agent number to snapshot metadata
# ---------------------------------------------------------------------------

@dataclass
class AgentSpec:
    number: int
    name: str
    output_type: Type[BaseModel]
    snapshot_file: str


AGENT_REGISTRY: dict[int, AgentSpec] = {
    1:  AgentSpec(1,  "research",    ResearchOutput,            "agent01_research.json"),
    2:  AgentSpec(2,  "analyst",     AnalystOutput,             "agent02_analyst.json"),
    3:  AgentSpec(3,  "rotation",    RotationDetectionOutput,   "agent03_rotation.json"),
    4:  AgentSpec(4,  "strategy",    SectorStrategyOutput,      "agent04_strategy.json"),
    5:  AgentSpec(5,  "portfolio",   PortfolioOutput,           "agent05_portfolio.json"),
    6:  AgentSpec(6,  "risk",        RiskAssessment,            "agent06_risk.json"),
    7:  AgentSpec(7,  "returns",     ReturnsProjectionOutput,   "agent07_returns.json"),
    8:  AgentSpec(8,  "reconciler",  ReconciliationOutput,      "agent08_reconciler.json"),
    9:  AgentSpec(9,  "educator",    EducatorOutput,            "agent09_educator.json"),
    10: AgentSpec(10, "synthesis",   SynthesisOutput,           "agent10_synthesis.json"),
    11: AgentSpec(11, "value",       ValueInvestorOutput,       "agent11_value.json"),
    12: AgentSpec(12, "pattern",     PortfolioPatternOutput,    "agent12_pattern.json"),
    13: AgentSpec(13, "historical",  HistoricalAnalysisOutput,  "agent13_historical.json"),
    14: AgentSpec(14, "target_return", TargetReturnOutput,      "agent14_target_return.json"),
    15: AgentSpec(15, "exit_strategy", ExitStrategyOutput,      "agent15_exit_strategy.json"),
    16: AgentSpec(16, "regime_detector", RegimeDetectorOutput,  "agent16_regime_detector.json"),
    17: AgentSpec(17, "earnings_risk", EarningsRiskOutput,      "agent17_earnings_risk.json"),
}


# ---------------------------------------------------------------------------
# Snapshot Save / Load
# ---------------------------------------------------------------------------

def _save_snapshot(output: BaseModel, agent_num: int, snapshots_dir: Path) -> Path:
    """Save agent output as JSON snapshot for incremental pipeline runs."""
    spec = AGENT_REGISTRY[agent_num]
    filepath = snapshots_dir / spec.snapshot_file
    filepath.write_text(output.model_dump_json(indent=2), encoding="utf-8")
    return filepath


def _load_snapshot(agent_num: int, snapshots_dir: Path) -> BaseModel:
    """Load agent output from JSON snapshot. Exits with error if missing."""
    spec = AGENT_REGISTRY[agent_num]
    filepath = snapshots_dir / spec.snapshot_file
    if not filepath.exists():
        print(f"\nERROR: Snapshot not found: {filepath}")
        print(f"Agent {spec.number:02d} ({spec.name}) output is required but no snapshot exists.")
        print(f"Run the full pipeline first, or run with --from {spec.number} or earlier.")
        sys.exit(1)
    json_str = filepath.read_text(encoding="utf-8")
    output = spec.output_type.model_validate_json(json_str)
    print(f"  [snapshot] Loaded agent {spec.number:02d} ({spec.name})")
    return output


# ---------------------------------------------------------------------------
# CLI Argument Parsing
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="IBD Momentum Investment Framework v4.0 — Multi-Agent Pipeline",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""\
examples:
  python run_pipeline.py                          full pipeline, default dirs
  python run_pipeline.py mydata results           custom data/output dirs
  python run_pipeline.py --from 5                 resume from Agent 05 (Portfolio)
  python run_pipeline.py --from 10 --historical   resume from Agent 10, include historical
""",
    )
    parser.add_argument(
        "data_dir_pos", nargs="?", default=None,
        help=argparse.SUPPRESS,  # hidden positional for backward compat
    )
    parser.add_argument(
        "output_dir_pos", nargs="?", default=None,
        help=argparse.SUPPRESS,
    )
    parser.add_argument(
        "--data", default="data",
        help="Data directory (default: data)",
    )
    parser.add_argument(
        "--output", default="output",
        help="Output directory (default: output)",
    )
    parser.add_argument(
        "--from", dest="start_from", type=int, default=1,
        choices=range(1, 18), metavar="N",
        help="Start from agent N, loading agents 1..N-1 from snapshots (1-16, default: 1)",
    )
    parser.add_argument(
        "--historical", action="store_true", default=False,
        help="Include Agent 13 (Historical Analyst). Off by default.",
    )

    args = parser.parse_args()

    # Positional args override defaults (backward compat)
    if args.data_dir_pos is not None:
        args.data = args.data_dir_pos
    if args.output_dir_pos is not None:
        args.output = args.output_dir_pos

    return args


def _write_research_excel(research_output, out_path: Path) -> Path:
    """Write Agent 01 Research output to its own Excel file."""
    today = date.today().isoformat()
    filepath = out_path / f"agent01_research_{today}.xlsx"

    # --- Stocks (split known sectors vs UNKNOWN) ---
    known_rows = []
    unknown_rows = []
    for s in research_output.stocks:
        row = {
            "Sector Rank": s.sector_rank,
            "Rank in Sector": s.stock_rank_in_sector,
            "Symbol": s.symbol,
            "Company": s.company_name,
            "Sector": s.sector,
            "Composite": s.composite_rating,
            "RS": s.rs_rating,
            "EPS": s.eps_rating,
            "SMR": s.smr_rating,
            "Acc/Dis": s.acc_dis_rating,
            "Tier": s.preliminary_tier,
            "Keep Candidate": s.is_ibd_keep_candidate,
            "Multi-Source": s.is_multi_source_validated,
            "Validation Score": s.validation_score,
            "Providers": s.validation_providers,
            "IBD Lists": ", ".join(s.ibd_lists),
            "Schwab Themes": ", ".join(s.schwab_themes),
            "Fool Status": s.fool_status,
            "M* Rating": s.morningstar_rating,
            "Moat": s.economic_moat,
            "Fair Value": s.fair_value,
            "M* Price": s.morningstar_price,
            "P/FV": s.price_to_fair_value,
            "Uncertainty": s.morningstar_uncertainty,
            "Confidence": s.confidence,
            "Reasoning": s.reasoning,
            "Sources": ", ".join(s.sources),
        }
        if s.sector == "UNKNOWN":
            unknown_rows.append(row)
        else:
            known_rows.append(row)

    df_stocks = pd.DataFrame(known_rows)
    if not df_stocks.empty:
        df_stocks = df_stocks.sort_values(
            by=["Sector", "Composite"],
            ascending=[True, False],
            na_position="last",
        ).reset_index(drop=True)

    df_unknown = pd.DataFrame(unknown_rows)

    # --- ETFs ---
    etf_rows = []
    for e in research_output.etfs:
        etf_rows.append({
            "Rank": e.etf_rank,
            "Symbol": e.symbol,
            "Name": e.name,
            "Tier": e.preliminary_tier,
            "RS Rating": e.rs_rating,
            "Acc/Dis": e.acc_dis_rating,
            "YTD Chg %": e.ytd_change,
            "Price Chg": e.price_change,
            "Vol % Chg": e.volume_pct_change,
            "Close": e.close_price,
            "Div Yield": e.div_yield,
            "ETF Score": e.etf_score,
            "Focus": e.focus,
            "Schwab Themes": ", ".join(e.schwab_themes),
            "IBD Lists": ", ".join(e.ibd_lists),
            "Key Theme ETF": e.key_theme_etf,
            "Sources": ", ".join(e.sources),
        })
    df_etfs = pd.DataFrame(etf_rows)
    if not df_etfs.empty:
        df_etfs = df_etfs.sort_values(
            by=["Rank"], na_position="last",
        ).reset_index(drop=True)

    # --- Sector Patterns ---
    sector_pattern_rows = []
    for sp in research_output.sector_patterns:
        sector_pattern_rows.append({
            "Sector": sp.sector,
            "Stock Count": sp.stock_count,
            "Avg Composite": sp.avg_composite,
            "Avg RS": sp.avg_rs,
            "Elite %": sp.elite_pct,
            "Multi-List %": sp.multi_list_pct,
            "Keep Count": sp.ibd_keep_count,
            "Strength": sp.strength,
            "Trend": sp.trend_direction,
            "Evidence": sp.evidence,
        })
    df_sector_patterns = pd.DataFrame(sector_pattern_rows)

    # --- Summary ---
    summary_rows = [
        {"Field": "Analysis Date", "Value": research_output.analysis_date},
        {"Field": "Total Stocks", "Value": len(research_output.stocks)},
        {"Field": "Total ETFs", "Value": len(research_output.etfs)},
        {"Field": "Sectors", "Value": len(research_output.sector_patterns)},
        {"Field": "IBD Keep Candidates", "Value": len(research_output.ibd_keep_candidates)},
        {"Field": "Multi-Source Validated", "Value": len(research_output.multi_source_validated)},
        {"Field": "Total Securities Scanned", "Value": research_output.total_securities_scanned},
        {"Field": "Data Sources Used", "Value": ", ".join(research_output.data_sources_used)},
        {"Field": "Data Sources Failed", "Value": ", ".join(research_output.data_sources_failed) or "None"},
        {"Field": "Summary", "Value": research_output.summary},
    ]
    df_summary = pd.DataFrame(summary_rows)

    # --- Write ---
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        df_stocks.to_excel(writer, sheet_name="Stocks", index=False)
        if not df_unknown.empty:
            df_unknown.to_excel(writer, sheet_name="Unknown Sector", index=False)
        if not df_etfs.empty:
            df_etfs.to_excel(writer, sheet_name="ETFs", index=False)
        if not df_sector_patterns.empty:
            df_sector_patterns.to_excel(writer, sheet_name="Sector Patterns", index=False)

    return filepath


def _write_analyst_excel(analyst_output, out_path: Path, research_output=None) -> Path:
    """Write Agent 02 Analyst output to its own Excel file."""
    today = date.today().isoformat()
    filepath = out_path / f"agent02_analyst_{today}.xlsx"

    # --- Top 100 + per-tier sheets ---
    def _elite_detail(passed, rating, threshold, label):
        """Format elite filter result: PASS/FAIL with rating vs threshold."""
        if passed is None:
            return f"N/A — {label} missing"
        elif passed:
            return f"PASS — {label} {rating} (>= {threshold})"
        else:
            return f"FAIL — {label} {rating} (< {threshold})"

    def _stock_row(s, rank):
        return {
            "Rank": rank,
            "Symbol": s.symbol,
            "Company": s.company_name,
            "Sector": s.sector,
            "Tier": s.tier,
            "Tier Label": s.tier_label,
            "Conviction": s.conviction,
            "Composite": s.composite_rating,
            "RS": s.rs_rating,
            "EPS": s.eps_rating,
            "SMR": s.smr_rating,
            "Acc/Dis": s.acc_dis_rating,
            "All Elite Pass": "PASS" if s.passes_all_elite else "FAIL",
            "EPS Elite (>=85)": _elite_detail(s.passes_eps_filter, s.eps_rating, 85, "EPS"),
            "RS Elite (>=75)": _elite_detail(s.passes_rs_filter, s.rs_rating, 75, "RS"),
            "SMR Elite (A/B/B-)": _elite_detail(s.passes_smr_filter, s.smr_rating, "A/B/B-", "SMR"),
            "AccDis Elite (A/B/B-)": _elite_detail(s.passes_acc_dis_filter, s.acc_dis_rating, "A/B/B-", "Acc/Dis"),
            "IBD Keep": s.is_ibd_keep,
            "Multi-Source": s.is_multi_source_validated,
            "Validation Score": s.validation_score,
            "Sector Rank": s.sector_rank_in_sector,
            "Est P/E": s.estimated_pe,
            "P/E Category": s.pe_category,
            "PEG Ratio": s.peg_ratio,
            "PEG Category": s.peg_category,
            "Est Beta": s.estimated_beta,
            "Est Return %": s.estimated_return_pct,
            "Return Cat": s.return_category,
            "Volatility %": s.estimated_volatility_pct,
            "Sharpe Ratio": s.sharpe_ratio,
            "Sharpe Cat": s.sharpe_category,
            "Alpha %": s.alpha_pct,
            "Alpha Cat": s.alpha_category,
            "Risk Rating": s.risk_rating,
            "LLM P/E": s.llm_pe_ratio,
            "LLM Fwd P/E": s.llm_forward_pe,
            "LLM Beta": s.llm_beta,
            "LLM 1Y Return %": s.llm_annual_return_1y,
            "LLM Volatility %": s.llm_volatility,
            "Div Yield %": s.llm_dividend_yield,
            "Mkt Cap ($B)": s.llm_market_cap_b,
            "LLM Valuation": s.llm_valuation_grade,
            "LLM Risk": s.llm_risk_grade,
            "Guidance": s.llm_guidance,
            "Data Source": s.valuation_source,
            "Strengths": "; ".join(s.strengths),
            "Weaknesses": "; ".join(s.weaknesses),
            "Catalyst": s.catalyst,
            "Catalyst Date": s.catalyst_date,
            "Catalyst Type": s.catalyst_type,
            "Catalyst Adj": s.catalyst_conviction_adjustment,
            "Catalyst Source": s.catalyst_source,
            "IBD Lists": ", ".join(s.ibd_lists),
            "Schwab Themes": ", ".join(s.schwab_themes),
            "M* Rating": s.morningstar_rating,
            "Moat": s.economic_moat,
            "Fair Value": s.fair_value,
            "M* Price": s.morningstar_price,
            "P/FV": s.price_to_fair_value,
            "Uncertainty": s.morningstar_uncertainty,
        }

    t1_rows, t2_rows, t3_rows = [], [], []
    for s in analyst_output.rated_stocks:
        if s.tier == 1:
            t1_rows.append(_stock_row(s, len(t1_rows) + 1))
        elif s.tier == 2:
            t2_rows.append(_stock_row(s, len(t2_rows) + 1))
        elif s.tier == 3:
            t3_rows.append(_stock_row(s, len(t3_rows) + 1))

    # Top 100 = first 100 across all tiers (already sorted by tier, conviction, score)
    top100_rows = []
    for i, s in enumerate(analyst_output.rated_stocks[:100], 1):
        top100_rows.append(_stock_row(s, i))

    df_top100 = pd.DataFrame(top100_rows)
    df_t1 = pd.DataFrame(t1_rows)
    df_t2 = pd.DataFrame(t2_rows)
    df_t3 = pd.DataFrame(t3_rows)

    # --- Sector Rankings ---
    sector_rank_rows = []
    for sr in analyst_output.sector_rankings:
        sector_rank_rows.append({
            "Rank": sr.rank,
            "Sector": sr.sector,
            "Score": sr.sector_score,
            "Stock Count": sr.stock_count,
            "Avg Composite": sr.avg_composite,
            "Avg RS": sr.avg_rs,
            "Elite %": sr.elite_pct,
            "Multi-List %": sr.multi_list_pct,
            "IBD Keeps": sr.ibd_keep_count,
            "Avg P/E": sr.avg_pe,
            "Avg Beta": sr.avg_beta,
            "Avg Volatility": sr.avg_volatility,
            "Top Stocks": ", ".join(sr.top_stocks),
        })
    df_sector_ranks = pd.DataFrame(sector_rank_rows)

    # --- IBD Keeps ---
    keep_rows = []
    for k in analyst_output.ibd_keeps:
        keep_rows.append({
            "Symbol": k.symbol,
            "Composite": k.composite_rating,
            "RS": k.rs_rating,
            "EPS": k.eps_rating,
            "Tier": k.tier,
            "IBD Lists": ", ".join(k.ibd_lists),
            "Rationale": k.keep_rationale,
            "Override Risk": k.override_risk or "",
        })
    df_keeps = pd.DataFrame(keep_rows)

    # --- Unrated ---
    unrated_rows = []
    for u in analyst_output.unrated_stocks:
        unrated_rows.append({
            "Symbol": u.symbol,
            "Company": u.company_name,
            "Sector": u.sector or "",
            "Reason": u.reason_unrated,
            "Schwab Themes": ", ".join(u.schwab_themes),
            "Validation Score": u.validation_score,
            "Note": u.note,
        })
    df_unrated = pd.DataFrame(unrated_rows)

    # --- Summary ---
    td = analyst_output.tier_distribution
    ef = analyst_output.elite_filter_summary
    summary_rows = [
        {"Field": "Analysis Date", "Value": analyst_output.analysis_date},
        {"Field": "Total Rated Stocks", "Value": len(analyst_output.rated_stocks)},
        {"Field": "Unrated Stocks", "Value": len(analyst_output.unrated_stocks)},
        {"Field": "IBD Keeps", "Value": len(analyst_output.ibd_keeps)},
        {"Field": "Sectors Ranked", "Value": len(analyst_output.sector_rankings)},
        {"Field": "", "Value": ""},
        {"Field": "--- Tier Distribution ---", "Value": ""},
        {"Field": "T1 Momentum", "Value": td.tier_1_count},
        {"Field": "T2 Quality Growth", "Value": td.tier_2_count},
        {"Field": "T3 Defensive", "Value": td.tier_3_count},
        {"Field": "Below Threshold", "Value": td.below_threshold_count},
        {"Field": "Unrated", "Value": td.unrated_count},
        {"Field": "", "Value": ""},
        {"Field": "--- Elite Screening ---", "Value": ""},
        {"Field": "Total Screened", "Value": ef.total_screened},
        {"Field": "Passed All 4", "Value": ef.passed_all_four},
        {"Field": "Failed EPS", "Value": ef.failed_eps},
        {"Field": "Failed RS", "Value": ef.failed_rs},
        {"Field": "Failed SMR", "Value": ef.failed_smr},
        {"Field": "Failed Acc/Dis", "Value": ef.failed_acc_dis},
        {"Field": "Missing Ratings", "Value": ef.missing_ratings},
    ]

    # Valuation summary sections
    vs = analyst_output.valuation_summary
    if vs:
        mc = vs.market_context
        summary_rows.extend([
            {"Field": "", "Value": ""},
            {"Field": "--- Market Context ---", "Value": ""},
            {"Field": "S&P 500 Forward P/E", "Value": f"{mc.sp500_forward_pe}x"},
            {"Field": "10-Year Average P/E", "Value": f"{mc.sp500_10y_avg_pe}x"},
            {"Field": "Current Premium", "Value": f"+{mc.current_premium_pct}%"},
            {"Field": "Risk-Free Rate", "Value": f"{mc.risk_free_rate}%"},
            {"Field": "Market Return Assumption", "Value": f"{mc.market_return_assumption}%"},
        ])

        # P/E Statistics by Tier
        summary_rows.extend([
            {"Field": "", "Value": ""},
            {"Field": "--- P/E Statistics by Tier ---", "Value": ""},
        ])
        for t in vs.tier_pe_stats:
            label = {1: "T1 Momentum", 2: "T2 Quality Growth", 3: "T3 Defensive"}.get(t.tier, f"T{t.tier}")
            summary_rows.extend([
                {"Field": f"{label} Avg P/E", "Value": f"{t.avg_pe:.1f}x"},
                {"Field": f"{label} Median P/E", "Value": f"{t.median_pe:.1f}x"},
                {"Field": f"{label} Range", "Value": f"{t.min_pe:.1f}x - {t.max_pe:.1f}x ({t.stock_count} stocks)"},
            ])

        # P/E Distribution
        pd_dist = vs.pe_distribution
        total = pd_dist.total or 1
        summary_rows.extend([
            {"Field": "", "Value": ""},
            {"Field": "--- P/E Distribution ---", "Value": ""},
            {"Field": "Deep Value", "Value": f"{pd_dist.deep_value_count} ({pd_dist.deep_value_count/total*100:.1f}%)"},
            {"Field": "Value", "Value": f"{pd_dist.value_count} ({pd_dist.value_count/total*100:.1f}%)"},
            {"Field": "Reasonable", "Value": f"{pd_dist.reasonable_count} ({pd_dist.reasonable_count/total*100:.1f}%)"},
            {"Field": "Growth Premium", "Value": f"{pd_dist.growth_premium_count} ({pd_dist.growth_premium_count/total*100:.1f}%)"},
            {"Field": "Speculative", "Value": f"{pd_dist.speculative_count} ({pd_dist.speculative_count/total*100:.1f}%)"},
        ])

        # Top 5 Sectors by P/E
        if vs.top_sectors_by_pe:
            summary_rows.extend([
                {"Field": "", "Value": ""},
                {"Field": "--- Top 5 Sectors by P/E ---", "Value": ""},
            ])
            for i, sec in enumerate(vs.top_sectors_by_pe, 1):
                summary_rows.append(
                    {"Field": f"#{i} {sec.sector}", "Value": f"{sec.avg_pe:.1f}x ({sec.stock_count} stocks)"}
                )

        # PEG Ratio Analysis
        peg = vs.peg_analysis
        summary_rows.extend([
            {"Field": "", "Value": ""},
            {"Field": "--- PEG Ratio Analysis ---", "Value": ""},
            {"Field": "Avg PEG", "Value": f"{peg.avg_peg:.2f}" if peg.avg_peg else "N/A"},
            {"Field": "Median PEG", "Value": f"{peg.median_peg:.2f}" if peg.median_peg else "N/A"},
            {"Field": "Undervalued (PEG<1.0)", "Value": f"{peg.undervalued_count} ({peg.pct_undervalued:.1f}%)"},
            {"Field": "Fair Value (PEG 1-2)", "Value": str(peg.fair_value_count)},
            {"Field": "Expensive (PEG>2)", "Value": str(peg.expensive_count)},
        ])

        # Portfolio Risk Metrics
        summary_rows.extend([
            {"Field": "", "Value": ""},
            {"Field": "--- Portfolio Risk Metrics ---", "Value": ""},
            {"Field": "Avg Sharpe Ratio", "Value": f"{vs.avg_sharpe:.2f}" if vs.avg_sharpe else "N/A"},
            {"Field": "Avg Beta", "Value": f"{vs.avg_beta:.2f}" if vs.avg_beta else "N/A"},
            {"Field": "Avg Alpha", "Value": f"{vs.avg_alpha:.1f}%" if vs.avg_alpha else "N/A"},
        ])

    # LLM Enrichment summary
    llm_count = sum(1 for s in analyst_output.rated_stocks if s.valuation_source == "llm")
    total_rated = len(analyst_output.rated_stocks)
    summary_rows.extend([
        {"Field": "", "Value": ""},
        {"Field": "--- LLM Enrichment ---", "Value": ""},
        {"Field": "Stocks Enriched", "Value": f"{llm_count} / {total_rated}"},
        {"Field": "Data Source", "Value": "Claude Haiku (training knowledge)" if llm_count > 0 else "Deterministic estimates only"},
        {"Field": "Note", "Value": "LLM metrics from training data (early 2025), not real-time" if llm_count > 0 else "No LLM enrichment — using sector-based estimates"},
    ])

    summary_rows.extend([
        {"Field": "", "Value": ""},
        {"Field": "Methodology", "Value": analyst_output.methodology_notes},
        {"Field": "Summary", "Value": analyst_output.summary},
    ])
    df_summary = pd.DataFrame(summary_rows)

    # --- Color fill definitions ---
    _DARK_GREEN = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    _LIGHT_GREEN = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    _YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    _ORANGE = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    _RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    _GREEN = PatternFill(start_color="00CC00", end_color="00CC00", fill_type="solid")
    _WHITE_FONT = Font(color="FFFFFF")

    # Map column header -> {cell value -> (fill, use_white_font)}
    _COLOR_MAP = {
        "P/E Category": {
            "Deep Value": (_DARK_GREEN, True),
            "Value": (_LIGHT_GREEN, False),
            "Reasonable": (_YELLOW, False),
            "Growth Premium": (_ORANGE, False),
            "Speculative": (_RED, True),
        },
        "PEG Category": {
            "Undervalued": (_GREEN, False),
            "Fair Value": (_YELLOW, False),
            "Expensive": (_ORANGE, False),
        },
        "Return Cat": {
            "Strong": (_GREEN, False),
            "Good": (_LIGHT_GREEN, False),
            "Moderate": (_YELLOW, False),
            "Weak": (_RED, True),
        },
        "Sharpe Cat": {
            "Excellent": (_DARK_GREEN, True),
            "Good": (_LIGHT_GREEN, False),
            "Moderate": (_YELLOW, False),
            "Below Average": (_ORANGE, False),
        },
        "Alpha Cat": {
            "Strong Outperformer": (_DARK_GREEN, True),
            "Outperformer": (_LIGHT_GREEN, False),
            "Slight Underperformer": (_YELLOW, False),
            "Underperformer": (_RED, True),
        },
        "Risk Rating": {
            "Excellent": (_DARK_GREEN, True),
            "Good": (_LIGHT_GREEN, False),
            "Moderate": (_YELLOW, False),
            "Below Average": (_ORANGE, False),
            "Poor": (_RED, True),
        },
        "LLM Valuation": {
            "Deep Value": (_DARK_GREEN, True),
            "Value": (_LIGHT_GREEN, False),
            "Reasonable": (_YELLOW, False),
            "Growth Premium": (_ORANGE, False),
            "Speculative": (_RED, True),
        },
        "LLM Risk": {
            "Excellent": (_DARK_GREEN, True),
            "Good": (_LIGHT_GREEN, False),
            "Moderate": (_YELLOW, False),
            "Below Average": (_ORANGE, False),
            "Poor": (_RED, True),
        },
    }

    def _apply_color_formatting(ws):
        """Apply color fills to category columns based on cell values."""
        # Build column letter map from header row
        header_map = {}
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col_idx).value
            if header in _COLOR_MAP:
                header_map[col_idx] = _COLOR_MAP[header]

        for row_idx in range(2, ws.max_row + 1):
            for col_idx, value_map in header_map.items():
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value in value_map:
                    fill, use_white = value_map[cell.value]
                    cell.fill = fill
                    if use_white:
                        cell.font = _WHITE_FONT

    # --- Rated ETFs ---
    etf_rows = []
    for e in analyst_output.rated_etfs:
        etf_rows.append({
            "Rank": e.etf_rank,
            "Symbol": e.symbol,
            "Name": e.name,
            "Focus": e.focus,
            "Focus Rank": e.focus_rank,
            "Tier": e.tier,
            "Conviction": e.conviction,
            "Screen": "PASS" if e.passes_etf_screen else "FAIL",
            "RS Rating": e.rs_rating,
            "Acc/Dis": e.acc_dis_rating,
            "YTD Chg %": e.ytd_change,
            "Vol % Chg": e.volume_pct_change,
            "Price Chg": e.price_change,
            "Close": e.close_price,
            "Div Yield": e.div_yield,
            "ETF Score": e.etf_score,
            "Themes": ", ".join(e.theme_tags),
            "Strengths": "; ".join(e.strengths),
            "Weaknesses": "; ".join(e.weaknesses),
        })
    df_etfs = pd.DataFrame(etf_rows)
    if not df_etfs.empty:
        df_etfs = df_etfs.sort_values(by=["Rank"], na_position="last").reset_index(drop=True)

    # --- Morningstar Analysis sheet ---
    # Morningstar Score: star_pts (0-40) + moat_pts (0-30) + discount_pts (0-20) + certainty_pts (0-10)
    _moat_pts = {"Wide": 30, "Narrow": 15, "None": 0}
    _unc_pts = {"Low": 10, "Medium": 7, "High": 3, "Very High": 0}
    ms_rows = []
    if research_output:
        for s in research_output.stocks:
            if s.morningstar_rating:
                # Star points: 5-star=40, 4-star=25
                star_pts = 40 if s.morningstar_rating == "5-star" else 25
                # Moat points: Wide=30, Narrow=15, None/missing=0
                moat_pts = _moat_pts.get(s.economic_moat, 0)
                # Discount points: based on P/FV (lower = more undervalued = higher score)
                # P/FV 0.50 -> 20pts, 0.70 -> 12pts, 0.90 -> 4pts, 1.00+ -> 0pts
                pfv = s.price_to_fair_value
                discount_pts = max(0, min(20, round((1.0 - pfv) * 40))) if pfv else 0
                # Certainty points: Low=10, Medium=7, High=3, Very High=0
                cert_pts = _unc_pts.get(s.morningstar_uncertainty, 0)
                ms_score = star_pts + moat_pts + discount_pts + cert_pts

                ms_rows.append({
                    "Rank": 0,  # filled after sort
                    "Symbol": s.symbol,
                    "Company": s.company_name,
                    "Sector": s.sector,
                    "M* Score": ms_score,
                    "M* Rating": s.morningstar_rating,
                    "Moat": s.economic_moat,
                    "Fair Value": s.fair_value,
                    "M* Price": s.morningstar_price,
                    "P/FV": s.price_to_fair_value,
                    "Discount %": round((1.0 - pfv) * 100, 1) if pfv else None,
                    "Uncertainty": s.morningstar_uncertainty,
                    "Composite": s.composite_rating,
                    "RS": s.rs_rating,
                    "EPS": s.eps_rating,
                    "Validation Score": s.validation_score,
                    "Multi-Source": s.is_multi_source_validated,
                    "Cap Size": s.cap_size,
                    "IBD Lists": ", ".join(s.ibd_lists) if s.ibd_lists else "",
                    "Schwab Themes": ", ".join(s.schwab_themes) if s.schwab_themes else "",
                    "Sources": ", ".join(s.sources),
                })
    df_morningstar = pd.DataFrame(ms_rows)
    if not df_morningstar.empty:
        df_morningstar = df_morningstar.sort_values(
            by=["M* Score"], ascending=False, na_position="last",
        ).reset_index(drop=True)
        df_morningstar["Rank"] = range(1, len(df_morningstar) + 1)

    # --- Write ---
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        df_top100.to_excel(writer, sheet_name="Top 100", index=False)
        if not df_t1.empty:
            df_t1.to_excel(writer, sheet_name="T1 Momentum", index=False)
        if not df_t2.empty:
            df_t2.to_excel(writer, sheet_name="T2 Quality Growth", index=False)
        if not df_t3.empty:
            df_t3.to_excel(writer, sheet_name="T3 Defensive", index=False)
        df_sector_ranks.to_excel(writer, sheet_name="Sector Rankings", index=False)
        if not df_keeps.empty:
            df_keeps.to_excel(writer, sheet_name="IBD Keeps", index=False)
        if not df_etfs.empty:
            df_etfs.to_excel(writer, sheet_name="Rated ETFs", index=False)
        if not df_morningstar.empty:
            df_morningstar.to_excel(writer, sheet_name="Morningstar", index=False)
        if not df_unrated.empty:
            df_unrated.to_excel(writer, sheet_name="Unrated", index=False)

        # Apply color formatting to stock sheets
        for sheet_name in ["Top 100", "T1 Momentum", "T2 Quality Growth", "T3 Defensive"]:
            if sheet_name in writer.book.sheetnames:
                _apply_color_formatting(writer.book[sheet_name])

    return filepath


def _write_rotation_excel(rotation_output, out_path: Path) -> Path:
    """Write Agent 03 Rotation Detector output to its own Excel file."""
    today = date.today().isoformat()
    filepath = out_path / f"agent03_rotation_{today}.xlsx"

    _DARK_GREEN = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    _LIGHT_GREEN = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    _YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    _RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    _WHITE_FONT = Font(color="FFFFFF")

    # --- Summary ---
    summary_rows = [
        {"Field": "Analysis Date", "Value": rotation_output.analysis_date},
        {"Field": "Verdict", "Value": rotation_output.verdict.value.upper()},
        {"Field": "Confidence", "Value": f"{rotation_output.confidence}%"},
        {"Field": "Rotation Type", "Value": rotation_output.rotation_type.value},
        {"Field": "Rotation Stage", "Value": rotation_output.rotation_stage.value if rotation_output.rotation_stage else "N/A"},
        {"Field": "Velocity", "Value": rotation_output.velocity or "N/A"},
        {"Field": "Regime", "Value": rotation_output.market_regime.regime},
        {"Field": "", "Value": ""},
        {"Field": "Signals Active", "Value": f"{rotation_output.signals.signals_active}/5"},
        {"Field": "Source Sectors", "Value": len(rotation_output.source_sectors)},
        {"Field": "Destination Sectors", "Value": len(rotation_output.destination_sectors)},
        {"Field": "Stable Sectors", "Value": len(rotation_output.stable_sectors)},
        {"Field": "", "Value": ""},
        {"Field": "Summary", "Value": rotation_output.summary},
        {"Field": "", "Value": ""},
        {"Field": "ETF Flow Summary", "Value": rotation_output.etf_flow_summary or "N/A"},
        {"Field": "", "Value": ""},
        {"Field": "Rotation Narrative", "Value": rotation_output.rotation_narrative or "N/A"},
        {"Field": "Narrative Source", "Value": rotation_output.narrative_source or "N/A"},
    ]
    df_summary = pd.DataFrame(summary_rows)

    # --- Signals ---
    signals = rotation_output.signals
    signal_rows = []
    for sig in [signals.rs_divergence, signals.leadership_change,
                signals.breadth_shift, signals.elite_concentration_shift,
                signals.ibd_keep_migration]:
        signal_rows.append({
            "Signal": sig.signal_name,
            "Triggered": "YES" if sig.triggered else "NO",
            "Value": sig.value,
            "Threshold": sig.threshold,
            "Evidence": sig.evidence,
            "ETF Confirmation": sig.etf_confirmation or "",
        })
    df_signals = pd.DataFrame(signal_rows)

    # --- Market Regime ---
    regime = rotation_output.market_regime
    regime_rows = [
        {"Field": "Regime", "Value": regime.regime},
        {"Field": "Sector Breadth %", "Value": regime.sector_breadth_pct},
        {"Field": "Bull Signals", "Value": "; ".join(regime.bull_signals_present) or "None"},
        {"Field": "Bear Signals", "Value": "; ".join(regime.bear_signals_present) or "None"},
        {"Field": "Regime Note", "Value": regime.regime_note},
    ]
    df_regime = pd.DataFrame(regime_rows)

    # --- Source Sectors (Outflow) ---
    source_rows = []
    for sf in rotation_output.source_sectors:
        source_rows.append({
            "Sector": sf.sector,
            "Cluster": sf.cluster,
            "Direction": sf.direction,
            "Rank": sf.current_rank,
            "Avg RS": sf.avg_rs,
            "Elite %": sf.elite_pct,
            "Stock Count": sf.stock_count,
            "Magnitude": sf.magnitude,
            "Evidence": sf.evidence,
        })
    df_source = pd.DataFrame(source_rows)

    # --- Destination Sectors (Inflow) ---
    dest_rows = []
    for sf in rotation_output.destination_sectors:
        dest_rows.append({
            "Sector": sf.sector,
            "Cluster": sf.cluster,
            "Direction": sf.direction,
            "Rank": sf.current_rank,
            "Avg RS": sf.avg_rs,
            "Elite %": sf.elite_pct,
            "Stock Count": sf.stock_count,
            "Magnitude": sf.magnitude,
            "Evidence": sf.evidence,
        })
    df_dest = pd.DataFrame(dest_rows)

    # --- Strategist Notes ---
    notes_rows = [{"#": i+1, "Note": note}
                  for i, note in enumerate(rotation_output.strategist_notes)]
    df_notes = pd.DataFrame(notes_rows)

    # --- Write ---
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        df_signals.to_excel(writer, sheet_name="Signals", index=False)
        df_regime.to_excel(writer, sheet_name="Market Regime", index=False)
        if not df_source.empty:
            df_source.to_excel(writer, sheet_name="Source Sectors", index=False)
        if not df_dest.empty:
            df_dest.to_excel(writer, sheet_name="Destination Sectors", index=False)
        df_notes.to_excel(writer, sheet_name="Strategist Notes", index=False)

        # Color-code signals: green=triggered, red=not
        ws_signals = writer.book["Signals"]
        for row_idx in range(2, ws_signals.max_row + 1):
            cell = ws_signals.cell(row=row_idx, column=2)  # Triggered column
            if cell.value == "YES":
                cell.fill = _LIGHT_GREEN
            elif cell.value == "NO":
                cell.fill = _RED
                cell.font = _WHITE_FONT

        # Color-code verdict on summary
        ws_summary = writer.book["Summary"]
        for row_idx in range(2, ws_summary.max_row + 1):
            field_cell = ws_summary.cell(row=row_idx, column=1)
            value_cell = ws_summary.cell(row=row_idx, column=2)
            if field_cell.value == "Verdict":
                if value_cell.value == "ACTIVE":
                    value_cell.fill = _RED
                    value_cell.font = _WHITE_FONT
                elif value_cell.value == "EMERGING":
                    value_cell.fill = _YELLOW
                elif value_cell.value == "NONE":
                    value_cell.fill = _LIGHT_GREEN

    return filepath


def _write_strategy_excel(strategy_output, out_path: Path) -> Path:
    """Write Agent 04 Sector Strategist output to its own Excel file."""
    today = date.today().isoformat()
    filepath = out_path / f"agent04_strategy_{today}.xlsx"

    _DARK_GREEN = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    _LIGHT_GREEN = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    _YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    _ORANGE = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    _WHITE_FONT = Font(color="FFFFFF")

    alloc = strategy_output.sector_allocations

    # --- Summary ---
    summary_rows = [
        {"Field": "Analysis Date", "Value": strategy_output.analysis_date},
        {"Field": "Rotation Response", "Value": strategy_output.rotation_response},
        {"Field": "Regime Adjustment", "Value": strategy_output.regime_adjustment},
        {"Field": "", "Value": ""},
        {"Field": "--- Tier Targets ---", "Value": ""},
        {"Field": "T1 (Momentum)", "Value": f"{alloc.tier_targets.get('T1', 0):.1f}%"},
        {"Field": "T2 (Quality)", "Value": f"{alloc.tier_targets.get('T2', 0):.1f}%"},
        {"Field": "T3 (Defensive)", "Value": f"{alloc.tier_targets.get('T3', 0):.1f}%"},
        {"Field": "Cash", "Value": f"{alloc.tier_targets.get('Cash', 0):.1f}%"},
        {"Field": "Cash Recommendation", "Value": f"{alloc.cash_recommendation:.1f}%"},
        {"Field": "", "Value": ""},
        {"Field": "Overall Sectors", "Value": len(alloc.overall_allocation)},
        {"Field": "Theme Recommendations", "Value": len(strategy_output.theme_recommendations)},
        {"Field": "Rotation Signals", "Value": len(strategy_output.rotation_signals)},
        {"Field": "", "Value": ""},
        {"Field": "Allocation Rationale", "Value": alloc.rationale},
        {"Field": "Summary", "Value": strategy_output.summary},
    ]
    df_summary = pd.DataFrame(summary_rows)

    # --- Overall Allocation ---
    overall_rows = [
        {"Sector": sector, "Allocation %": pct}
        for sector, pct in sorted(alloc.overall_allocation.items(), key=lambda x: -x[1])
    ]
    df_overall = pd.DataFrame(overall_rows)

    # --- T1 Allocation ---
    t1_rows = [
        {"Sector": sector, "Allocation %": pct}
        for sector, pct in sorted(alloc.tier_1_allocation.items(), key=lambda x: -x[1])
    ]
    df_t1 = pd.DataFrame(t1_rows)

    # --- T2 Allocation ---
    t2_rows = [
        {"Sector": sector, "Allocation %": pct}
        for sector, pct in sorted(alloc.tier_2_allocation.items(), key=lambda x: -x[1])
    ]
    df_t2 = pd.DataFrame(t2_rows)

    # --- T3 Allocation ---
    t3_rows = [
        {"Sector": sector, "Allocation %": pct}
        for sector, pct in sorted(alloc.tier_3_allocation.items(), key=lambda x: -x[1])
    ]
    df_t3 = pd.DataFrame(t3_rows)

    # --- Theme Recommendations ---
    theme_rows = []
    for rec in strategy_output.theme_recommendations:
        etf_rank_str = ""
        if rec.etf_rankings:
            etf_rank_str = "; ".join(
                f"{r.get('symbol', '?')}(score={r.get('etf_score', '?')}, rank={r.get('rank', '?')})"
                for r in rec.etf_rankings
            )
        theme_rows.append({
            "Theme": rec.theme,
            "ETFs": ", ".join(rec.recommended_etfs),
            "Tier Fit": f"T{rec.tier_fit}",
            "Allocation": rec.allocation_suggestion,
            "Conviction": rec.conviction,
            "Rationale": rec.rationale,
            "ETF Rankings": etf_rank_str,
        })
    df_themes = pd.DataFrame(theme_rows)

    # --- Rotation Signals ---
    signal_rows = []
    for sig in strategy_output.rotation_signals:
        signal_rows.append({
            "Action": sig.action,
            "Trigger": sig.trigger,
            "Confirmation": "; ".join(sig.confirmation),
            "Invalidation": "; ".join(sig.invalidation),
        })
    df_signals = pd.DataFrame(signal_rows)

    # --- Write ---
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        df_overall.to_excel(writer, sheet_name="Overall Allocation", index=False)
        df_t1.to_excel(writer, sheet_name="T1 Allocation", index=False)
        df_t2.to_excel(writer, sheet_name="T2 Allocation", index=False)
        df_t3.to_excel(writer, sheet_name="T3 Allocation", index=False)
        if not df_themes.empty:
            df_themes.to_excel(writer, sheet_name="Theme Recommendations", index=False)
        if not df_signals.empty:
            df_signals.to_excel(writer, sheet_name="Rotation Signals", index=False)

        # Color-code conviction in Theme Recommendations
        if "Theme Recommendations" in writer.book.sheetnames:
            ws = writer.book["Theme Recommendations"]
            # Find conviction column
            conv_col = None
            for col_idx in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col_idx).value == "Conviction":
                    conv_col = col_idx
                    break
            if conv_col:
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=conv_col)
                    if cell.value == "HIGH":
                        cell.fill = _LIGHT_GREEN
                    elif cell.value == "MEDIUM":
                        cell.fill = _YELLOW
                    elif cell.value == "LOW":
                        cell.fill = _ORANGE

    return filepath


def _write_portfolio_excel(portfolio_output, out_path: Path) -> Path:
    """Write Agent 05 Portfolio Manager output to its own Excel file."""
    today = date.today().isoformat()
    filepath = out_path / f"agent05_portfolio_{today}.xlsx"

    # --- Summary ---
    t1 = portfolio_output.tier_1
    t2 = portfolio_output.tier_2
    t3 = portfolio_output.tier_3
    summary_rows = [
        {"Field": "Analysis Date", "Value": portfolio_output.analysis_date},
        {"Field": "Total Positions", "Value": portfolio_output.total_positions},
        {"Field": "Stock Count", "Value": portfolio_output.stock_count},
        {"Field": "ETF Count", "Value": portfolio_output.etf_count},
        {"Field": "", "Value": ""},
        {"Field": "--- Tier Allocation ---", "Value": ""},
        {"Field": "T1 Momentum", "Value": f"{t1.actual_pct:.1f}% ({len(t1.positions)} pos)"},
        {"Field": "T2 Quality Growth", "Value": f"{t2.actual_pct:.1f}% ({len(t2.positions)} pos)"},
        {"Field": "T3 Defensive", "Value": f"{t3.actual_pct:.1f}% ({len(t3.positions)} pos)"},
        {"Field": "Cash", "Value": f"{portfolio_output.cash_pct:.1f}%"},
        {"Field": "", "Value": ""},
        {"Field": "--- Keeps ---", "Value": ""},
        {"Field": "Fundamental Keeps", "Value": len(portfolio_output.keeps_placement.fundamental_keeps)},
        {"Field": "IBD Keeps", "Value": len(portfolio_output.keeps_placement.ibd_keeps)},
        {"Field": "Total Keeps", "Value": portfolio_output.keeps_placement.total_keeps},
        {"Field": "", "Value": ""},
        {"Field": "Methodology", "Value": portfolio_output.construction_methodology},
        {"Field": "Summary", "Value": portfolio_output.summary},
    ]
    df_summary = pd.DataFrame(summary_rows)

    # --- Positions per tier ---
    def _pos_row(p, rank):
        return {
            "Rank": rank,
            "Symbol": p.symbol,
            "Company": p.company_name,
            "Sector": p.sector,
            "Type": p.asset_type,
            "Target %": p.target_pct,
            "Stop %": p.trailing_stop_pct,
            "Max Loss %": p.max_loss_pct,
            "Keep": p.keep_category or "",
            "Conviction": p.conviction,
            "Vol Adj %": p.volatility_adjustment,
            "Sizing Source": p.sizing_source or "",
            "Reasoning": p.reasoning,
        }

    df_t1 = pd.DataFrame([_pos_row(p, i+1) for i, p in enumerate(t1.positions)])
    df_t2 = pd.DataFrame([_pos_row(p, i+1) for i, p in enumerate(t2.positions)])
    df_t3 = pd.DataFrame([_pos_row(p, i+1) for i, p in enumerate(t3.positions)])

    # --- Sector Exposure ---
    sector_rows = [
        {"Sector": sec, "Allocation %": pct}
        for sec, pct in sorted(portfolio_output.sector_exposure.items(), key=lambda x: -x[1])
    ]
    df_sectors = pd.DataFrame(sector_rows)

    # --- Orders ---
    order_rows = []
    for o in portfolio_output.orders:
        order_rows.append({
            "Symbol": o.symbol,
            "Action": o.action,
            "Tier": o.tier,
            "Target %": o.target_pct,
            "Rationale": o.rationale,
        })
    df_orders = pd.DataFrame(order_rows)

    # --- Write ---
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        if not df_t1.empty:
            df_t1.to_excel(writer, sheet_name="T1 Momentum", index=False)
        if not df_t2.empty:
            df_t2.to_excel(writer, sheet_name="T2 Quality Growth", index=False)
        if not df_t3.empty:
            df_t3.to_excel(writer, sheet_name="T3 Defensive", index=False)
        if not df_sectors.empty:
            df_sectors.to_excel(writer, sheet_name="Sector Exposure", index=False)
        if not df_orders.empty:
            df_orders.to_excel(writer, sheet_name="Orders", index=False)

    return filepath


def _write_risk_excel(risk_output, out_path: Path) -> Path:
    """Write Agent 06 Risk Officer output to its own Excel file."""
    today = date.today().isoformat()
    filepath = out_path / f"agent06_risk_{today}.xlsx"

    # --- Summary ---
    swl = risk_output.sleep_well_scores
    summary_rows = [
        {"Field": "Analysis Date", "Value": risk_output.analysis_date},
        {"Field": "Overall Status", "Value": risk_output.overall_status},
        {"Field": "Checks Run", "Value": len(risk_output.check_results)},
        {"Field": "Vetoes", "Value": len(risk_output.vetoes)},
        {"Field": "Warnings", "Value": len(risk_output.warnings)},
        {"Field": "", "Value": ""},
        {"Field": "--- Sleep Well Scores ---", "Value": ""},
        {"Field": "Overall", "Value": f"{swl.overall_score}/10"},
        {"Field": "T1 Score", "Value": f"{swl.tier_1_score}/10"},
        {"Field": "T2 Score", "Value": f"{swl.tier_2_score}/10"},
        {"Field": "T3 Score", "Value": f"{swl.tier_3_score}/10"},
        {"Field": "Factors", "Value": "; ".join(swl.factors)},
        {"Field": "", "Value": ""},
        {"Field": "Stop-Loss Source", "Value": risk_output.stop_loss_source or "N/A"},
        {"Field": "Stop-Loss Recommendations", "Value": len(risk_output.stop_loss_recommendations)},
        {"Field": "", "Value": ""},
        {"Field": "Summary", "Value": risk_output.summary},
    ]
    df_summary = pd.DataFrame(summary_rows)

    # --- Check Results ---
    check_rows = []
    for c in risk_output.check_results:
        check_rows.append({
            "Check": c.check_name,
            "Status": c.status,
            "Findings": c.findings,
            "Details": "; ".join(c.details) if c.details else "",
        })
    df_checks = pd.DataFrame(check_rows)

    # --- Stress Tests ---
    stress_rows = []
    for s in risk_output.stress_test_results.scenarios:
        stress_rows.append({
            "Scenario": s.scenario_name,
            "Impact": s.impact_description,
            "Est Drawdown %": s.estimated_drawdown_pct,
            "Most Affected": ", ".join(s.positions_most_affected[:5]),
        })
    df_stress = pd.DataFrame(stress_rows)

    # --- Stop-Loss Recommendations ---
    stop_rows = []
    for rec in risk_output.stop_loss_recommendations:
        stop_rows.append({
            "Symbol": rec.symbol,
            "Current Stop %": rec.current_stop_pct,
            "Recommended Stop %": rec.recommended_stop_pct,
            "Reason": rec.reason,
            "Volatility": rec.volatility_flag or "",
        })
    df_stops = pd.DataFrame(stop_rows)

    # --- Write ---
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        df_checks.to_excel(writer, sheet_name="Check Results", index=False)
        if not df_stress.empty:
            df_stress.to_excel(writer, sheet_name="Stress Tests", index=False)
        if not df_stops.empty:
            df_stops.to_excel(writer, sheet_name="Stop-Loss Recommendations", index=False)

    return filepath


def _write_regime_detector_excel(regime_output, out_path: Path) -> Path:
    """Write Agent 16 Regime Detector output to its own Excel file."""
    today = date.today().isoformat()
    filepath = out_path / f"agent16_regime_detector_{today}.xlsx"

    # --- Regime Summary ---
    summary_rows = [
        {"Field": "Analysis Date", "Value": regime_output.analysis_date},
        {"Field": "Market Regime", "Value": regime_output.regime.value},
        {"Field": "Legacy Regime", "Value": regime_output.to_legacy_regime()},
        {"Field": "Confidence", "Value": regime_output.confidence.value},
        {"Field": "Market Health Score", "Value": f"{regime_output.market_health_score}/10"},
        {"Field": "Exposure Recommendation", "Value": f"{regime_output.exposure_recommendation}%"},
        {"Field": "", "Value": ""},
        {"Field": "Bullish Signals", "Value": regime_output.bullish_signals},
        {"Field": "Neutral Signals", "Value": regime_output.neutral_signals},
        {"Field": "Bearish Signals", "Value": regime_output.bearish_signals},
        {"Field": "", "Value": ""},
        {"Field": "Executive Summary", "Value": regime_output.executive_summary},
    ]
    if regime_output.regime_change:
        rc = regime_output.regime_change
        summary_rows.extend([
            {"Field": "", "Value": ""},
            {"Field": "--- Regime Change ---", "Value": ""},
            {"Field": "Changed", "Value": str(rc.changed)},
            {"Field": "Previous", "Value": rc.previous},
            {"Field": "Current", "Value": rc.current},
            {"Field": "Trigger", "Value": rc.trigger or "N/A"},
        ])
    df_summary = pd.DataFrame(summary_rows)

    # --- Indicator Assessments ---
    assessment_rows = []

    # Distribution days
    dist = regime_output.distribution_days
    assessment_rows.append({
        "Indicator": "Distribution Days",
        "Signal": dist.signal.value,
        "Detail": dist.detail,
        "SP500 Count": dist.sp500_count,
        "Nasdaq Count": dist.nasdaq_count,
        "SP500 Direction": dist.sp500_direction.value,
        "Nasdaq Direction": dist.nasdaq_direction.value,
    })

    # Breadth
    b = regime_output.breadth
    assessment_rows.append({
        "Indicator": "Market Breadth",
        "Signal": b.signal.value,
        "Detail": f"200MA: {b.pct_above_200ma}%, 50MA: {b.pct_above_50ma}%, "
                  f"Highs/Lows: {b.new_highs}/{b.new_lows}",
        "SP500 Count": "",
        "Nasdaq Count": "",
        "SP500 Direction": b.advance_decline_direction.value,
        "Nasdaq Direction": "",
    })

    # Leaders
    ldr = regime_output.leaders
    assessment_rows.append({
        "Indicator": "Leading Stocks",
        "Signal": ldr.signal.value,
        "Detail": f"RS90 above 50MA: {ldr.rs90_above_50ma_pct}%, "
                  f"Health: {ldr.health.value}, Highs/Lows: {ldr.rs90_new_highs}/{ldr.rs90_new_lows}",
        "SP500 Count": "",
        "Nasdaq Count": "",
        "SP500 Direction": "",
        "Nasdaq Direction": "",
    })

    # Sectors
    sec = regime_output.sectors
    assessment_rows.append({
        "Indicator": "Sector Rotation",
        "Signal": sec.signal.value,
        "Detail": f"Character: {sec.character.value}, "
                  f"Defensive in top 5: {sec.defensive_in_top_5}, "
                  f"Top: {', '.join(sec.top_5_sectors[:3])}",
        "SP500 Count": "",
        "Nasdaq Count": "",
        "SP500 Direction": "",
        "Nasdaq Direction": "",
    })

    # FTD
    ftd = regime_output.follow_through_day
    if ftd.detected:
        ftd_signal = "BULLISH" if ftd.quality.value in ("STRONG", "MODERATE") else "NEUTRAL"
        ftd_detail = (f"Quality: {ftd.quality.value}, "
                      f"Index: {ftd.index or 'N/A'}, Gain: {ftd.gain_pct or 0:.2f}%, "
                      f"Day {ftd.rally_day_number or 'N/A'}")
    else:
        ftd_signal = "NEUTRAL"
        ftd_detail = "No Follow-Through Day detected"
    assessment_rows.append({
        "Indicator": "Follow-Through Day",
        "Signal": ftd_signal,
        "Detail": ftd_detail,
        "SP500 Count": "",
        "Nasdaq Count": "",
        "SP500 Direction": "",
        "Nasdaq Direction": "",
    })

    df_assessments = pd.DataFrame(assessment_rows)

    # --- Transition Conditions ---
    trans_rows = []
    for tc in regime_output.transition_conditions:
        trans_rows.append({
            "Direction": tc.direction,
            "Target Regime": tc.target_regime,
            "Condition": tc.condition,
            "Likelihood": tc.likelihood or "N/A",
        })
    df_transitions = pd.DataFrame(trans_rows)

    # --- Write ---
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Regime Summary", index=False)
        df_assessments.to_excel(writer, sheet_name="Indicator Assessments", index=False)
        if not df_transitions.empty:
            df_transitions.to_excel(writer, sheet_name="Transition Conditions", index=False)

        # Color-code regime in summary
        ws = writer.sheets["Regime Summary"]
        regime_fills = {
            "CONFIRMED_UPTREND": PatternFill(start_color="228B22", end_color="228B22", fill_type="solid"),
            "UPTREND_UNDER_PRESSURE": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
            "FOLLOW_THROUGH_DAY": PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid"),
            "RALLY_ATTEMPT": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
            "CORRECTION": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
        }
        # Color the regime value cell (row 2, column B)
        regime_cell = ws.cell(row=2, column=2)
        fill = regime_fills.get(regime_cell.value)
        if fill:
            regime_cell.fill = fill
            regime_cell.font = Font(bold=True, color="FFFFFF" if regime_cell.value in ("CONFIRMED_UPTREND", "CORRECTION") else "000000")

        # Color-code signal cells in assessments
        ws_assess = writer.sheets["Indicator Assessments"]
        signal_fills = {
            "BULLISH": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
            "NEUTRAL": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
            "BEARISH": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        }
        signal_col = 2  # Column B = Signal (1-indexed)
        for row_idx in range(2, ws_assess.max_row + 1):
            cell = ws_assess.cell(row=row_idx, column=signal_col)
            fill = signal_fills.get(cell.value)
            if fill:
                cell.fill = fill

    return filepath


def _write_exit_strategy_excel(exit_output, out_path: Path) -> Path:
    """Write Agent 15 Exit Strategist output to its own Excel file."""
    today = date.today().isoformat()
    filepath = out_path / f"agent15_exit_strategy_{today}.xlsx"

    # --- Summary ---
    impact = exit_output.portfolio_impact
    summary_rows = [
        {"Field": "Analysis Date", "Value": exit_output.analysis_date},
        {"Field": "Market Regime", "Value": exit_output.market_regime.value},
        {"Field": "Portfolio Health Score", "Value": f"{exit_output.portfolio_health_score}/10"},
        {"Field": "Total Signals", "Value": len(exit_output.signals)},
        {"Field": "Critical", "Value": impact.positions_critical},
        {"Field": "Warning", "Value": impact.positions_warning},
        {"Field": "Watch", "Value": impact.positions_watch},
        {"Field": "Healthy", "Value": impact.positions_healthy},
        {"Field": "", "Value": ""},
        {"Field": "Current Cash %", "Value": f"{impact.current_cash_pct:.1f}%"},
        {"Field": "Projected Cash %", "Value": f"{impact.projected_cash_pct:.1f}%"},
        {"Field": "Sector Concentration Risk", "Value": impact.sector_concentration_risk},
        {"Field": "Reasoning Source", "Value": exit_output.reasoning_source or "N/A"},
        {"Field": "", "Value": ""},
        {"Field": "Summary", "Value": exit_output.summary},
    ]
    df_summary = pd.DataFrame(summary_rows)

    # --- Signals ---
    signal_rows = []
    for s in exit_output.signals:
        signal_rows.append({
            "Symbol": s.symbol,
            "Tier": s.tier,
            "Type": s.asset_type,
            "Urgency": s.urgency.value,
            "Action": s.action.value,
            "Sell Type": s.sell_type.value,
            "Sell %": f"{s.sell_pct:.0f}%" if s.sell_pct is not None else "",
            "Gain/Loss %": f"{s.gain_loss_pct:+.1f}%",
            "Current Price": f"${s.current_price:.2f}",
            "Buy Price": f"${s.buy_price:.2f}",
            "Stop Price": f"${s.stop_price:.2f}",
            "Days Held": s.days_held,
            "Rules Triggered": ", ".join(r.value for r in s.rules_triggered),
            "Reasoning": s.reasoning,
        })
    df_signals = pd.DataFrame(signal_rows)

    # --- Evidence ---
    evidence_rows = []
    for s in exit_output.signals:
        for e in s.evidence:
            evidence_rows.append({
                "Symbol": s.symbol,
                "Rule": e.rule_triggered.value,
                "Rule ID": e.rule_id,
                "Data Point": e.data_point,
            })
    df_evidence = pd.DataFrame(evidence_rows)

    # --- Portfolio Impact ---
    impact_rows = [
        {"Metric": "Current Cash %", "Value": f"{impact.current_cash_pct:.1f}%"},
        {"Metric": "Projected Cash %", "Value": f"{impact.projected_cash_pct:.1f}%"},
        {"Metric": "Current Top Holding %", "Value": f"{impact.current_top_holding_pct:.1f}%"},
        {"Metric": "Projected Top Holding %", "Value": f"{impact.projected_top_holding_pct:.1f}%"},
        {"Metric": "Positions Healthy", "Value": impact.positions_healthy},
        {"Metric": "Positions Watch", "Value": impact.positions_watch},
        {"Metric": "Positions Warning", "Value": impact.positions_warning},
        {"Metric": "Positions Critical", "Value": impact.positions_critical},
        {"Metric": "Sector Concentration Risk", "Value": impact.sector_concentration_risk},
    ]
    df_impact = pd.DataFrame(impact_rows)

    # --- Rules Summary ---
    from ibd_agents.schemas.exit_strategy_output import SELL_RULE_NAMES, RULE_ID_MAP
    rules_rows = []
    for rule_name in SELL_RULE_NAMES:
        triggered_count = sum(
            1 for s in exit_output.signals
            for r in s.rules_triggered if r.value == rule_name
        )
        rules_rows.append({
            "Rule": rule_name,
            "Rule ID": RULE_ID_MAP.get(rule_name, ""),
            "Triggered Count": triggered_count,
        })
    df_rules = pd.DataFrame(rules_rows)

    # --- Write ---
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        df_signals.to_excel(writer, sheet_name="Signals", index=False)
        if not df_evidence.empty:
            df_evidence.to_excel(writer, sheet_name="Evidence", index=False)
        df_impact.to_excel(writer, sheet_name="Portfolio Impact", index=False)
        df_rules.to_excel(writer, sheet_name="Rules Summary", index=False)

        # Color-code signals by urgency
        ws = writer.sheets["Signals"]
        urgency_fills = {
            "CRITICAL": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
            "WARNING": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
            "WATCH": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
            "HEALTHY": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
        }
        urgency_col = 4  # Column D = Urgency (1-indexed)
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=urgency_col)
            fill = urgency_fills.get(cell.value)
            if fill:
                cell.fill = fill

    return filepath


def _write_earnings_risk_excel(earnings_output, out_path: Path) -> Path:
    """Write Agent 17 Earnings Risk Analyst output to its own Excel file."""
    today = date.today().isoformat()
    filepath = out_path / f"agent17_earnings_risk_{today}.xlsx"

    # --- Summary ---
    conc = earnings_output.concentration
    summary_rows = [
        {"Field": "Analysis Date", "Value": earnings_output.analysis_date},
        {"Field": "Market Regime", "Value": earnings_output.market_regime},
        {"Field": "Lookforward Days", "Value": earnings_output.lookforward_days},
        {"Field": "Data Source", "Value": earnings_output.data_source or "N/A"},
        {"Field": "", "Value": ""},
        {"Field": "Positions with Earnings", "Value": len(earnings_output.analyses)},
        {"Field": "Positions Clear", "Value": len(earnings_output.positions_clear)},
        {"Field": "Concentration Risk", "Value": conc.concentration_risk},
        {"Field": "Portfolio % Exposed", "Value": f"{conc.total_portfolio_pct_exposed:.1f}%"},
        {"Field": "", "Value": ""},
        {"Field": "Executive Summary", "Value": earnings_output.executive_summary},
    ]
    df_summary = pd.DataFrame(summary_rows)

    # --- Analyses ---
    analysis_rows = []
    for a in earnings_output.analyses:
        analysis_rows.append({
            "Ticker": a.ticker,
            "Account": a.account,
            "Earnings Date": a.earnings_date,
            "Days Until": a.days_until_earnings,
            "Reporting Time": a.reporting_time,
            "Shares": a.shares,
            "Price": f"${a.current_price:.2f}",
            "Buy Price": f"${a.buy_price:.2f}",
            "Gain/Loss %": f"{a.gain_loss_pct:+.1f}%",
            "Position Value": f"${a.position_value:,.0f}",
            "Portfolio %": f"{a.portfolio_pct:.1f}%",
            "Cushion Ratio": f"{a.cushion_ratio:.2f}",
            "Cushion Cat.": a.cushion_category.value,
            "Risk Level": a.risk_level.value,
            "Beat Rate": f"{a.historical.beat_rate_pct:.0f}%",
            "Avg Move %": f"{a.historical.avg_move_pct:.1f}%",
            "Worst Move %": f"{a.historical.max_adverse_move_pct:.1f}%",
            "Estimate Rev.": a.estimate_revision.value,
            "Recommended": a.recommended_strategy.value,
            "Rationale": a.recommendation_rationale,
        })
    df_analyses = pd.DataFrame(analysis_rows)

    # --- Strategies ---
    strat_rows = []
    for a in earnings_output.analyses:
        for s in a.strategies:
            for sc in s.scenarios:
                strat_rows.append({
                    "Ticker": a.ticker,
                    "Strategy": s.strategy.value,
                    "Description": s.description,
                    "Shares to Sell": s.shares_to_sell or "",
                    "Hedge Cost": f"${s.estimated_hedge_cost:,.0f}" if s.estimated_hedge_cost else "",
                    "Scenario": sc.scenario,
                    "Move %": f"{sc.expected_move_pct:+.1f}%",
                    "Result G/L %": f"{sc.resulting_gain_loss_pct:+.1f}%",
                    "$ Impact": f"${sc.dollar_impact:+,.0f}",
                    "Math": sc.math,
                    "Risk/Reward": s.risk_reward_summary,
                })
    df_strategies = pd.DataFrame(strat_rows)

    # --- Concentration ---
    conc_rows = []
    for week in conc.earnings_calendar:
        conc_rows.append({
            "Week": week.week_label,
            "Week Start": week.week_start,
            "Positions": ", ".join(week.positions_reporting),
            "Portfolio %": f"{week.aggregate_portfolio_pct:.1f}%",
            "Flag": week.concentration_flag or "",
        })
    df_concentration = pd.DataFrame(conc_rows)

    # --- Write ---
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        if not df_analyses.empty:
            df_analyses.to_excel(writer, sheet_name="Analyses", index=False)
        if not df_strategies.empty:
            df_strategies.to_excel(writer, sheet_name="Strategies", index=False)
        if not df_concentration.empty:
            df_concentration.to_excel(writer, sheet_name="Concentration", index=False)

        # Color-code analyses by risk level
        if not df_analyses.empty:
            ws = writer.sheets["Analyses"]
            risk_fills = {
                "CRITICAL": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
                "HIGH": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
                "MODERATE": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
                "LOW": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
            }
            risk_col = 14  # Column N = Risk Level (1-indexed)
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=risk_col)
                fill = risk_fills.get(cell.value)
                if fill:
                    cell.fill = fill

    return filepath


def _write_returns_projector_excel(returns_output, out_path: Path) -> Path:
    """Write Agent 07 Returns Projector output to its own Excel file."""
    today = date.today().isoformat()
    filepath = out_path / f"agent07_returns_{today}.xlsx"

    ta = returns_output.tier_allocation
    er = returns_output.expected_return
    rm = returns_output.risk_metrics
    ci = returns_output.confidence_intervals

    # --- Summary ---
    summary_rows = [
        {"Field": "Analysis Date", "Value": returns_output.analysis_date},
        {"Field": "Portfolio Value", "Value": f"${returns_output.portfolio_value:,.0f}"},
        {"Field": "Market Regime", "Value": returns_output.market_regime},
        {"Field": "Regime Source", "Value": returns_output.regime_source},
        {"Field": "", "Value": ""},
        {"Field": "--- Tier Allocation ---", "Value": ""},
        {"Field": "Tier 1 %", "Value": f"{ta.tier_1_pct:.1f}%"},
        {"Field": "Tier 2 %", "Value": f"{ta.tier_2_pct:.1f}%"},
        {"Field": "Tier 3 %", "Value": f"{ta.tier_3_pct:.1f}%"},
        {"Field": "Cash %", "Value": f"{ta.cash_pct:.1f}%"},
        {"Field": "", "Value": ""},
        {"Field": "--- Expected Returns ---", "Value": ""},
        {"Field": "Expected 3m", "Value": f"{er.expected_3m:.2f}%"},
        {"Field": "Expected 6m", "Value": f"{er.expected_6m:.2f}%"},
        {"Field": "Expected 12m", "Value": f"{er.expected_12m:.2f}%"},
        {"Field": "", "Value": ""},
        {"Field": "--- Risk Metrics ---", "Value": ""},
        {"Field": "Max DD (with stops)", "Value": f"{rm.max_drawdown_with_stops:.2f}%"},
        {"Field": "Max DD (no stops)", "Value": f"{rm.max_drawdown_without_stops:.2f}%"},
        {"Field": "Max DD ($)", "Value": f"${rm.max_drawdown_dollar:,.0f}"},
        {"Field": "Projected Volatility", "Value": f"{rm.projected_annual_volatility:.2f}%"},
        {"Field": "Sharpe (base)", "Value": f"{rm.portfolio_sharpe_base:.2f}"},
        {"Field": "Return/Risk", "Value": f"{rm.return_per_unit_risk:.2f}"},
        {"Field": "", "Value": ""},
        {"Field": "Summary", "Value": returns_output.summary},
    ]
    df_summary = pd.DataFrame(summary_rows)

    # --- Scenarios ---
    scenario_rows = []
    for s in returns_output.scenarios:
        scenario_rows.append({
            "Scenario": s.scenario.title(),
            "Probability": f"{s.probability:.0%}",
            "T1 Return": f"{s.tier_1_return_pct:.1f}%",
            "T2 Return": f"{s.tier_2_return_pct:.1f}%",
            "T3 Return": f"{s.tier_3_return_pct:.1f}%",
            "Portfolio 3m": f"{s.portfolio_return_3m:.2f}%",
            "Portfolio 6m": f"{s.portfolio_return_6m:.2f}%",
            "Portfolio 12m": f"{s.portfolio_return_12m:.2f}%",
            "12m Gain ($)": f"${s.portfolio_gain_12m:,.0f}",
            "Ending Value": f"${s.ending_value_12m:,.0f}",
            "Reasoning": s.reasoning,
        })
    df_scenarios = pd.DataFrame(scenario_rows)

    # --- Benchmarks ---
    bench_rows = []
    for b in returns_output.benchmark_comparisons:
        bench_rows.append({
            "Benchmark": f"{b.symbol} ({b.name})",
            "Return 3m": f"{b.benchmark_return_3m:.2f}%",
            "Return 6m": f"{b.benchmark_return_6m:.2f}%",
            "Return 12m": f"{b.benchmark_return_12m:.2f}%",
            "Alpha (Bull)": f"{b.alpha_bull_12m:.2f}%",
            "Alpha (Base)": f"{b.alpha_base_12m:.2f}%",
            "Alpha (Bear)": f"{b.alpha_bear_12m:.2f}%",
            "Expected Alpha": f"{b.expected_alpha_12m:.2f}%",
            "Outperform Prob": f"{b.outperform_probability:.0%}",
        })
    df_benchmarks = pd.DataFrame(bench_rows)

    # --- Tier Mix Alternatives ---
    mix_rows = []
    for m in returns_output.tier_mix_comparison:
        mix_rows.append({
            "Mix": m.label,
            "T1 %": f"{m.tier_1_pct:.0f}%",
            "T2 %": f"{m.tier_2_pct:.0f}%",
            "T3 %": f"{m.tier_3_pct:.0f}%",
            "Cash %": f"{m.cash_pct:.0f}%",
            "Expected 12m": f"{m.expected_return_12m:.2f}%",
            "Bull 12m": f"{m.bull_return_12m:.2f}%",
            "Bear 12m": f"{m.bear_return_12m:.2f}%",
            "Max DD (stops)": f"{m.max_drawdown_with_stops:.2f}%",
            "Sharpe (base)": f"{m.sharpe_ratio_base:.2f}",
            "Note": m.comparison_note,
        })
    df_mixes = pd.DataFrame(mix_rows)

    # --- Confidence Intervals ---
    ci_rows = []
    for label, pr in [("3-Month", ci.horizon_3m), ("6-Month", ci.horizon_6m), ("12-Month", ci.horizon_12m)]:
        ci_rows.append({
            "Horizon": label,
            "P10": f"{pr.p10:.2f}%",
            "P25": f"{pr.p25:.2f}%",
            "P50 (Median)": f"{pr.p50:.2f}%",
            "P75": f"{pr.p75:.2f}%",
            "P90": f"{pr.p90:.2f}%",
            "P10 ($)": f"${pr.p10_dollar:,.0f}",
            "P50 ($)": f"${pr.p50_dollar:,.0f}",
            "P90 ($)": f"${pr.p90_dollar:,.0f}",
        })
    df_ci = pd.DataFrame(ci_rows)

    # --- Caveats ---
    df_caveats = pd.DataFrame({"Caveat": returns_output.caveats})

    # --- Asset Decomposition ---
    decomp_rows = []
    if returns_output.asset_type_decomposition:
        for d in returns_output.asset_type_decomposition:
            decomp_rows.append({
                "Tier": d.tier,
                "Stock %": f"{d.stock_contribution_pct:.2f}%",
                "ETF %": f"{d.etf_contribution_pct:.2f}%",
                "Stock Count": d.stock_count,
                "ETF Count": d.etf_count,
                "Stock Vol": f"{d.stock_avg_volatility:.2f}%" if d.stock_avg_volatility is not None else "N/A",
                "ETF Vol": f"{d.etf_avg_volatility:.2f}%" if d.etf_avg_volatility is not None else "N/A",
            })
    df_decomp = pd.DataFrame(decomp_rows)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        df_scenarios.to_excel(writer, sheet_name="Scenarios", index=False)
        df_benchmarks.to_excel(writer, sheet_name="Benchmarks", index=False)
        df_mixes.to_excel(writer, sheet_name="Tier Mix Alternatives", index=False)
        df_ci.to_excel(writer, sheet_name="Confidence Intervals", index=False)
        if not df_decomp.empty:
            df_decomp.to_excel(writer, sheet_name="Asset Decomposition", index=False)
        df_caveats.to_excel(writer, sheet_name="Caveats", index=False)

    return filepath


def _write_reconciler_excel(reconciler_output, out_path: Path) -> Path:
    """Write Agent 08 Reconciler output to its own Excel file."""
    today = date.today().isoformat()
    filepath = out_path / f"agent08_reconciler_{today}.xlsx"

    mf = reconciler_output.money_flow
    metrics = reconciler_output.transformation_metrics
    kv = reconciler_output.keep_verification

    # --- Summary ---
    summary_rows = [
        {"Field": "Analysis Date", "Value": reconciler_output.analysis_date},
        {"Field": "Total Actions", "Value": len(reconciler_output.actions)},
        {"Field": "Current Holdings", "Value": len(reconciler_output.current_holdings.holdings)},
        {"Field": "Total Value", "Value": f"${reconciler_output.current_holdings.total_value:,.0f}"},
        {"Field": "", "Value": ""},
        {"Field": "--- Money Flow ---", "Value": ""},
        {"Field": "Sell Proceeds", "Value": f"${mf.sell_proceeds:,.0f}"},
        {"Field": "Trim Proceeds", "Value": f"${mf.trim_proceeds:,.0f}"},
        {"Field": "Buy Cost", "Value": f"${mf.buy_cost:,.0f}"},
        {"Field": "Add Cost", "Value": f"${mf.add_cost:,.0f}"},
        {"Field": "Net Cash Change", "Value": f"${mf.net_cash_change:,.0f}"},
        {"Field": "Cash Reserve %", "Value": f"{mf.cash_reserve_pct:.1f}%"},
        {"Field": "", "Value": ""},
        {"Field": "--- Transformation ---", "Value": ""},
        {"Field": "Before Positions", "Value": metrics.before_position_count},
        {"Field": "After Positions", "Value": metrics.after_position_count},
        {"Field": "Before Sectors", "Value": metrics.before_sector_count},
        {"Field": "After Sectors", "Value": metrics.after_sector_count},
        {"Field": "Turnover %", "Value": f"{metrics.turnover_pct:.0f}%"},
        {"Field": "", "Value": ""},
        {"Field": "--- Keep Verification ---", "Value": ""},
        {"Field": "Keeps in Current", "Value": len(kv.keeps_in_current)},
        {"Field": "Keeps Missing", "Value": ", ".join(kv.keeps_missing) or "None"},
        {"Field": "Keeps to Buy", "Value": ", ".join(kv.keeps_to_buy) or "None"},
        {"Field": "", "Value": ""},
        {"Field": "Summary", "Value": reconciler_output.summary},
        {"Field": "", "Value": ""},
        {"Field": "Rationale Source", "Value": reconciler_output.rationale_source or "N/A"},
    ]
    df_summary = pd.DataFrame(summary_rows)

    # --- Actions ---
    action_rows = []
    for a in reconciler_output.actions:
        action_rows.append({
            "Symbol": a.symbol,
            "Action": a.action_type,
            "Current %": a.current_pct,
            "Target %": a.target_pct,
            "Dollar Change": f"${a.dollar_change:,.0f}",
            "Priority": a.priority,
            "Week": a.week,
            "Rationale": a.rationale,
        })
    df_actions = pd.DataFrame(action_rows)

    # --- Implementation Plan ---
    plan_rows = []
    for w in reconciler_output.implementation_plan:
        plan_rows.append({
            "Week": w.week,
            "Phase": w.phase_name,
            "Actions": len(w.actions),
        })
    df_plan = pd.DataFrame(plan_rows)

    # --- ETF Implementation Notes ---
    etf_notes_rows = [{"Note": note} for note in reconciler_output.etf_implementation_notes]
    df_etf_notes = pd.DataFrame(etf_notes_rows) if etf_notes_rows else pd.DataFrame()

    # --- Write ---
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        if not df_actions.empty:
            df_actions.to_excel(writer, sheet_name="Actions", index=False)
        df_plan.to_excel(writer, sheet_name="Implementation Plan", index=False)
        if not df_etf_notes.empty:
            df_etf_notes.to_excel(writer, sheet_name="ETF Implementation Notes", index=False)

    return filepath


def _write_educator_excel(educator_output, out_path: Path) -> Path:
    """Write Agent 09 Educator output to its own Excel file."""
    today = date.today().isoformat()
    filepath = out_path / f"agent09_educator_{today}.xlsx"

    # --- Summary ---
    summary_rows = [
        {"Field": "Analysis Date", "Value": educator_output.analysis_date},
        {"Field": "Stocks Explained", "Value": len(educator_output.stock_explanations)},
        {"Field": "Concept Lessons", "Value": len(educator_output.concept_lessons)},
        {"Field": "Glossary Entries", "Value": len(educator_output.glossary)},
        {"Field": "Action Items", "Value": len(educator_output.action_items)},
        {"Field": "", "Value": ""},
        {"Field": "Executive Summary", "Value": educator_output.executive_summary},
        {"Field": "", "Value": ""},
        {"Field": "Rotation Explanation", "Value": educator_output.rotation_explanation},
        {"Field": "", "Value": ""},
        {"Field": "Risk Explainer", "Value": educator_output.risk_explainer},
        {"Field": "", "Value": ""},
        {"Field": "Summary", "Value": educator_output.summary},
    ]
    df_summary = pd.DataFrame(summary_rows)

    # --- Stock Explanations ---
    stock_rows = []
    for se in educator_output.stock_explanations:
        stock_rows.append({
            "Symbol": se.symbol,
            "Company": se.company_name,
            "Tier": se.tier,
            "Keep": se.keep_category or "",
            "One-Liner": se.one_liner,
            "Why Selected": se.why_selected,
            "IBD Ratings": se.ibd_ratings_explained,
            "Strength": se.key_strength,
            "Risk": se.key_risk,
            "Position Context": se.position_context,
        })
    df_stocks = pd.DataFrame(stock_rows)

    # --- Concept Lessons ---
    lesson_rows = []
    for cl in educator_output.concept_lessons:
        lesson_rows.append({
            "Concept": cl.concept,
            "Explanation": cl.simple_explanation,
            "Analogy": cl.analogy,
            "Why It Matters": cl.why_it_matters,
            "Example": cl.example_from_analysis,
            "Reference": cl.framework_reference,
        })
    df_lessons = pd.DataFrame(lesson_rows)

    # --- Action Items ---
    action_rows = [{"#": i+1, "Action Item": item}
                   for i, item in enumerate(educator_output.action_items)]
    df_actions = pd.DataFrame(action_rows)

    # --- Glossary ---
    glossary_rows = [{"Term": term, "Definition": defn}
                     for term, defn in sorted(educator_output.glossary.items())]
    df_glossary = pd.DataFrame(glossary_rows)

    # --- ETF Explanations ---
    etf_expl_rows = []
    for ee in educator_output.etf_explanations:
        etf_expl_rows.append({
            "Symbol": ee.symbol,
            "Name": ee.name,
            "Tier": ee.tier,
            "One-Liner": ee.one_liner,
            "Why Selected": ee.why_selected,
            "Theme Context": ee.theme_context or "",
            "Conviction": ee.conviction_explained or "",
            "Key Strength": ee.key_strength,
            "Key Risk": ee.key_risk,
        })
    df_etf_expl = pd.DataFrame(etf_expl_rows) if etf_expl_rows else pd.DataFrame()

    # --- Write ---
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        if not df_stocks.empty:
            df_stocks.to_excel(writer, sheet_name="Stock Explanations", index=False)
        if not df_etf_expl.empty:
            df_etf_expl.to_excel(writer, sheet_name="ETF Explanations", index=False)
        if not df_lessons.empty:
            df_lessons.to_excel(writer, sheet_name="Concept Lessons", index=False)
        if not df_actions.empty:
            df_actions.to_excel(writer, sheet_name="Action Items", index=False)
        if not df_glossary.empty:
            df_glossary.to_excel(writer, sheet_name="Glossary", index=False)

    return filepath


def _write_synthesizer_excel(synthesis_output, out_path: Path) -> Path:
    """Write Agent 10 Executive Summary Synthesizer output to Excel."""
    today = date.today().isoformat()
    filepath = out_path / f"agent10_executive_summary_{today}.xlsx"

    # --- Investment Thesis ---
    thesis_rows = [
        {"Section": "Investment Thesis", "Content": synthesis_output.investment_thesis},
        {"Section": "", "Content": ""},
        {"Section": "Portfolio Narrative", "Content": synthesis_output.portfolio_narrative},
        {"Section": "", "Content": ""},
        {"Section": "Risk/Reward Assessment", "Content": synthesis_output.risk_reward_assessment},
        {"Section": "", "Content": ""},
        {"Section": "Market Context", "Content": synthesis_output.market_context},
    ]
    df_thesis = pd.DataFrame(thesis_rows)

    # --- Key Numbers Dashboard ---
    kn = synthesis_output.key_numbers
    numbers_rows = [
        {"Category": "Research", "Metric": "Total Securities Scanned", "Value": kn.total_stocks_scanned},
        {"Category": "Research", "Metric": "Stocks in Universe", "Value": kn.stocks_in_universe},
        {"Category": "Research", "Metric": "ETFs in Universe", "Value": kn.etfs_in_universe},
        {"Category": "Analyst", "Metric": "Stocks Rated", "Value": kn.stocks_rated},
        {"Category": "Analyst", "Metric": "ETFs Rated", "Value": kn.etfs_rated},
        {"Category": "Analyst", "Metric": "Tier 1 Count", "Value": kn.tier_1_count},
        {"Category": "Analyst", "Metric": "Tier 2 Count", "Value": kn.tier_2_count},
        {"Category": "Analyst", "Metric": "Tier 3 Count", "Value": kn.tier_3_count},
        {"Category": "Rotation", "Metric": "Verdict", "Value": kn.rotation_verdict},
        {"Category": "Rotation", "Metric": "Confidence", "Value": f"{kn.rotation_confidence}%"},
        {"Category": "Rotation", "Metric": "Market Regime", "Value": kn.market_regime},
        {"Category": "Portfolio", "Metric": "Total Positions", "Value": kn.portfolio_positions},
        {"Category": "Portfolio", "Metric": "Stock Count", "Value": kn.stock_count},
        {"Category": "Portfolio", "Metric": "ETF Count", "Value": kn.etf_count},
        {"Category": "Portfolio", "Metric": "Cash %", "Value": f"{kn.cash_pct:.1f}%"},
        {"Category": "Risk", "Metric": "Status", "Value": kn.risk_status},
        {"Category": "Risk", "Metric": "Sleep Well Score", "Value": f"{kn.sleep_well_score}/10"},
        {"Category": "Risk", "Metric": "Warnings", "Value": kn.risk_warnings_count},
        {"Category": "Returns", "Metric": "Bull Case 12m", "Value": f"{kn.bull_case_return_12m:.1f}%"},
        {"Category": "Returns", "Metric": "Base Case 12m", "Value": f"{kn.base_case_return_12m:.1f}%"},
        {"Category": "Returns", "Metric": "Bear Case 12m", "Value": f"{kn.bear_case_return_12m:.1f}%"},
        {"Category": "Reconciler", "Metric": "Turnover", "Value": f"{kn.turnover_pct:.0f}%"},
        {"Category": "Reconciler", "Metric": "Actions Count", "Value": kn.actions_count},
    ]
    df_numbers = pd.DataFrame(numbers_rows)

    # --- Cross-Agent Connections ---
    conn_rows = []
    for conn in synthesis_output.cross_agent_connections:
        conn_rows.append({
            "From Agent": conn.from_agent,
            "To Agent": conn.to_agent,
            "Connection": conn.connection,
            "Implication": conn.implication,
        })
    df_connections = pd.DataFrame(conn_rows)

    # --- Contradictions ---
    contra_rows = []
    for c in synthesis_output.contradictions:
        contra_rows.append({
            "Agent A": c.agent_a,
            "Agent B": c.agent_b,
            "Finding A": c.finding_a,
            "Finding B": c.finding_b,
            "Resolution": c.resolution,
        })
    df_contradictions = pd.DataFrame(contra_rows) if contra_rows else pd.DataFrame(
        [{"Agent A": "—", "Agent B": "—", "Finding A": "No contradictions detected",
          "Finding B": "—", "Resolution": "—"}]
    )

    # --- Action Items ---
    action_rows = [{"#": i+1, "Action Item": item}
                   for i, item in enumerate(synthesis_output.action_items)]
    df_actions = pd.DataFrame(action_rows)

    # --- Summary ---
    summary_rows = [
        {"Field": "Synthesis Source", "Value": synthesis_output.synthesis_source},
        {"Field": "Analysis Date", "Value": synthesis_output.analysis_date},
        {"Field": "Connections", "Value": len(synthesis_output.cross_agent_connections)},
        {"Field": "Contradictions", "Value": len(synthesis_output.contradictions)},
        {"Field": "Action Items", "Value": len(synthesis_output.action_items)},
        {"Field": "", "Value": ""},
        {"Field": "Summary", "Value": synthesis_output.summary},
    ]
    df_summary = pd.DataFrame(summary_rows)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df_thesis.to_excel(writer, sheet_name="Investment Thesis", index=False)
        df_numbers.to_excel(writer, sheet_name="Key Numbers", index=False)
        df_connections.to_excel(writer, sheet_name="Cross-Agent Connections", index=False)
        df_contradictions.to_excel(writer, sheet_name="Contradictions", index=False)
        df_actions.to_excel(writer, sheet_name="Action Items", index=False)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)

        # Format column widths
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 80)

    return filepath


# ---------------------------------------------------------------------------
# Agent 11: Value Investor Excel
# ---------------------------------------------------------------------------

_VALUE_CAT_COLORS = {
    "Quality Value": "006400",   # dark green
    "Deep Value": "228B22",      # forest green
    "GARP": "FFD700",            # gold
    "Dividend Value": "FF8C00",  # dark orange
    "Not Value": None,           # no fill
}

_TRAP_COLORS = {
    "High": "FF0000",     # red
    "Moderate": "FFA500",  # orange
    "Low": "FFFF00",       # yellow
    "None": None,
}


def _write_value_investor_excel(value_output, out_path: Path) -> Path:
    """Write Agent 11 Value Investor output to Excel."""
    today = date.today().isoformat()
    filepath = out_path / f"agent11_value_investor_{today}.xlsx"

    # --- Sheet 1: Summary ---
    cat_counts = {}
    for vs in value_output.value_stocks:
        cat_counts[vs.value_category] = cat_counts.get(vs.value_category, 0) + 1

    summary_rows = [
        {"Field": "Analysis Date", "Value": value_output.analysis_date},
        {"Field": "Total Stocks Analyzed", "Value": len(value_output.value_stocks)},
        {"Field": "--- Categories ---", "Value": ""},
    ]
    for cat in ("Quality Value", "Deep Value", "GARP", "Dividend Value", "Not Value"):
        summary_rows.append({"Field": f"  {cat}", "Value": cat_counts.get(cat, 0)})
    summary_rows.extend([
        {"Field": "--- Alerts ---", "Value": ""},
        {"Field": "Value Traps", "Value": len(value_output.value_traps)},
        {"Field": "M-V Mismatches", "Value": len(value_output.momentum_value_mismatches)},
        {"Field": "Top Value Picks", "Value": len(value_output.top_value_picks)},
        {"Field": "--- Moat ---", "Value": ""},
        {"Field": "Wide Moat", "Value": value_output.moat_analysis.wide_moat_count},
        {"Field": "Narrow Moat", "Value": value_output.moat_analysis.narrow_moat_count},
        {"Field": "No Moat", "Value": value_output.moat_analysis.no_moat_count},
        {"Field": "No Data", "Value": value_output.moat_analysis.no_data_count},
        {"Field": "--- Methodology ---", "Value": ""},
        {"Field": "Notes", "Value": value_output.methodology_notes},
    ])
    df_summary = pd.DataFrame(summary_rows)

    # --- Sheet 2: Value Rankings ---
    ranking_rows = []
    for vs in value_output.value_stocks:
        ranking_rows.append({
            "Rank": vs.value_rank,
            "Symbol": vs.symbol,
            "Company": vs.company_name,
            "Sector": vs.sector,
            "Tier": vs.tier,
            "Value Score": vs.value_score,
            "Category": vs.value_category,
            "M* Score": vs.morningstar_score,
            "PE Score": vs.pe_value_score,
            "PEG Score": vs.peg_value_score,
            "Moat Score": vs.moat_quality_score,
            "Discount Score": vs.discount_score,
            "PE Category": vs.pe_category,
            "PEG Ratio": vs.peg_ratio,
            "P/FV": vs.price_to_fair_value,
            "Moat": vs.economic_moat,
            "M* Rating": vs.morningstar_rating,
            "Fair Value": vs.fair_value,
            "Div Yield %": vs.llm_dividend_yield,
            "Composite": vs.composite_rating,
            "RS": vs.rs_rating,
            "EPS": vs.eps_rating,
            "Trap Risk": vs.value_trap_risk_level,
            "MV Alignment": vs.momentum_value_alignment,
            "Reasoning": vs.value_category_reasoning,
        })
    df_rankings = pd.DataFrame(ranking_rows)

    # --- Sheet 3: Value Categories ---
    cat_rows = []
    for cs in value_output.value_category_summaries:
        cat_rows.append({
            "Category": cs.category,
            "Count": cs.stock_count,
            "Avg Value Score": cs.avg_value_score,
            "Avg P/E": cs.avg_pe,
            "Avg PEG": cs.avg_peg,
            "Avg P/FV": cs.avg_pfv,
            "Top Stocks": ", ".join(cs.top_stocks[:5]),
            "Description": cs.description,
        })
    df_categories = pd.DataFrame(cat_rows)

    # --- Sheet 4: Moat Analysis ---
    moat = value_output.moat_analysis
    moat_rows = [
        {"Moat Type": "Wide", "Count": moat.wide_moat_count,
         "Avg Value Score": moat.avg_value_score_wide,
         "Avg P/FV": moat.avg_pfv_wide,
         "Stocks": ", ".join(moat.wide_moat_stocks[:10])},
        {"Moat Type": "Narrow", "Count": moat.narrow_moat_count,
         "Avg Value Score": moat.avg_value_score_narrow,
         "Avg P/FV": moat.avg_pfv_narrow,
         "Stocks": ""},
        {"Moat Type": "None", "Count": moat.no_moat_count,
         "Avg Value Score": None, "Avg P/FV": None, "Stocks": ""},
        {"Moat Type": "No Data", "Count": moat.no_data_count,
         "Avg Value Score": None, "Avg P/FV": None, "Stocks": ""},
    ]
    df_moat = pd.DataFrame(moat_rows)

    # --- Sheet 5: Value Traps ---
    trap_rows = []
    for vs in value_output.value_stocks:
        if vs.is_value_trap_risk:
            trap_rows.append({
                "Symbol": vs.symbol,
                "Company": vs.company_name,
                "Sector": vs.sector,
                "Value Score": vs.value_score,
                "Category": vs.value_category,
                "Trap Risk": vs.value_trap_risk_level,
                "Signals": "; ".join(vs.value_trap_signals),
                "RS": vs.rs_rating,
                "EPS": vs.eps_rating,
                "SMR": vs.smr_rating,
                "Acc/Dis": vs.acc_dis_rating,
            })
    df_traps = pd.DataFrame(trap_rows) if trap_rows else pd.DataFrame(
        columns=["Symbol", "Company", "Sector", "Value Score", "Category",
                 "Trap Risk", "Signals", "RS", "EPS", "SMR", "Acc/Dis"]
    )

    # --- Write Excel ---
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        df_rankings.to_excel(writer, sheet_name="Value Rankings", index=False)
        df_categories.to_excel(writer, sheet_name="Categories", index=False)
        df_moat.to_excel(writer, sheet_name="Moat Analysis", index=False)
        df_traps.to_excel(writer, sheet_name="Value Traps", index=False)

        # Color-code Category column in Rankings sheet
        ws = writer.book["Value Rankings"]
        cat_col = 7  # Column G = Category
        trap_col = 23  # Column W = Trap Risk
        for row_idx in range(2, len(ranking_rows) + 2):
            # Category color
            cat_val = ws.cell(row=row_idx, column=cat_col).value
            color = _VALUE_CAT_COLORS.get(cat_val)
            if color:
                ws.cell(row=row_idx, column=cat_col).fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
                ws.cell(row=row_idx, column=cat_col).font = Font(
                    color="FFFFFF" if color in ("006400", "228B22") else "000000"
                )
            # Trap Risk color
            trap_val = ws.cell(row=row_idx, column=trap_col).value
            trap_color = _TRAP_COLORS.get(trap_val)
            if trap_color:
                ws.cell(row=row_idx, column=trap_col).fill = PatternFill(
                    start_color=trap_color, end_color=trap_color, fill_type="solid"
                )

    return filepath


def _write_pattern_excel(pattern_output, out_path: Path) -> Path:
    """Write Agent 12 PatternAlpha per-stock scoring output to Excel."""
    today = date.today().isoformat()
    filepath = out_path / f"agent12_pattern_alpha_{today}.xlsx"

    ps = pattern_output.portfolio_summary

    # --- Sheet 1: Summary ---
    summary_rows = [
        {"Field": "Analysis Date", "Value": pattern_output.analysis_date},
        {"Field": "Scoring Source", "Value": pattern_output.scoring_source},
        {"Field": "", "Value": ""},
        {"Field": "--- Portfolio Summary ---", "Value": ""},
        {"Field": "Total Stocks Scored", "Value": ps.total_stocks_scored},
        {"Field": "Stocks With Patterns", "Value": ps.stocks_with_patterns},
        {"Field": "Avg Enhanced Score", "Value": f"{ps.avg_enhanced_score:.1f}/150"},
        {"Field": "Tier 1 Candidates", "Value": ps.tier_1_candidates},
        {"Field": "Category Kings (P4>=8)", "Value": ps.category_kings},
        {"Field": "Inflection Alerts (P5>=6)", "Value": ps.inflection_alerts},
        {"Field": "Disruption Risks (P2=0)", "Value": ps.disruption_risks},
        {"Field": "", "Value": ""},
        {"Field": "Methodology", "Value": pattern_output.methodology_notes},
        {"Field": "Summary", "Value": pattern_output.summary},
    ]
    df_summary = pd.DataFrame(summary_rows)

    # --- Sheet 2: Enhanced Rankings ---
    ranking_rows = []
    sorted_analyses = sorted(
        pattern_output.stock_analyses,
        key=lambda sa: sa.enhanced_score,
        reverse=True,
    )
    for rank, sa in enumerate(sorted_analyses, 1):
        pattern_total = sa.pattern_score.pattern_total if sa.pattern_score else 0
        dominant = sa.pattern_score.dominant_pattern if sa.pattern_score else ""
        ranking_rows.append({
            "Rank": rank,
            "Symbol": sa.symbol,
            "Company": sa.company_name,
            "Sector": sa.sector,
            "Tier": sa.tier,
            "Base (100)": sa.base_score.base_total,
            "Pattern (50)": pattern_total,
            "Enhanced (150)": sa.enhanced_score,
            "Stars": sa.enhanced_rating,
            "Rating": sa.enhanced_rating_label,
            "Dominant Pattern": dominant,
            "Source": sa.scoring_source,
        })
    df_rankings = pd.DataFrame(ranking_rows)

    # --- Sheet 3: Pattern Details ---
    detail_rows = []
    for sa in sorted_analyses:
        if sa.pattern_score is None:
            continue
        ps_b = sa.pattern_score
        detail_rows.append({
            "Symbol": sa.symbol,
            "Company": sa.company_name,
            "P1 Platform (12)": ps_b.p1_platform.score,
            "P2 Cannibalize (10)": ps_b.p2_cannibalization.score,
            "P3 Capital (10)": ps_b.p3_capital_allocation.score,
            "P4 Category (10)": ps_b.p4_category_creation.score,
            "P5 Inflection (8)": ps_b.p5_inflection_timing.score,
            "Pattern Total": ps_b.pattern_total,
            "Dominant": ps_b.dominant_pattern,
            "P1 Justification": ps_b.p1_platform.justification,
            "P2 Justification": ps_b.p2_cannibalization.justification,
            "P3 Justification": ps_b.p3_capital_allocation.justification,
            "P4 Justification": ps_b.p4_category_creation.justification,
            "P5 Justification": ps_b.p5_inflection_timing.justification,
            "Pattern Narrative": ps_b.pattern_narrative,
        })
    df_details = pd.DataFrame(detail_rows) if detail_rows else pd.DataFrame()

    # --- Sheet 4: Pattern Alerts ---
    alert_rows = []
    for alert in pattern_output.pattern_alerts:
        alert_rows.append({
            "Alert Type": alert.alert_type,
            "Symbol": alert.symbol,
            "Pattern": alert.pattern_name,
            "Score": alert.pattern_score,
            "Description": alert.description,
        })
    df_alerts = pd.DataFrame(alert_rows) if alert_rows else pd.DataFrame()

    # --- Sheet 5: Tier Assessment ---
    tier_rows = []
    for sa in sorted_analyses:
        tier_rows.append({
            "Symbol": sa.symbol,
            "Tier": sa.tier,
            "Enhanced": sa.enhanced_score,
            "Stars": sa.enhanced_rating,
            "Tier Recommendation": sa.tier_recommendation,
            "Alerts": ", ".join(sa.pattern_alerts) if sa.pattern_alerts else "",
        })
    df_tiers = pd.DataFrame(tier_rows)

    # Color definitions
    _GREEN = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    _LIGHT_GREEN = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    _YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    _ORANGE = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    _RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    _WHITE_FONT = Font(color="FFFFFF")

    # Rating label to color mapping
    _RATING_COLORS = {
        "Max Conviction": _GREEN,
        "Strong": _LIGHT_GREEN,
        "Favorable": _YELLOW,
        "Notable": _ORANGE,
        "Monitor": _RED,
        "Review": _RED,
    }

    # Alert type to color mapping
    _ALERT_COLORS = {
        "Category King": _GREEN,
        "Inflection Alert": _YELLOW,
        "Disruption Risk": _RED,
        "Pattern Imbalance": _ORANGE,
    }

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        df_rankings.to_excel(writer, sheet_name="Enhanced Rankings", index=False)
        if not df_details.empty:
            df_details.to_excel(writer, sheet_name="Pattern Details", index=False)
        if not df_alerts.empty:
            df_alerts.to_excel(writer, sheet_name="Pattern Alerts", index=False)
        df_tiers.to_excel(writer, sheet_name="Tier Assessment", index=False)

        # Color-code enhanced rankings by rating label (column J = 10)
        ws_rank = writer.book["Enhanced Rankings"]
        for row_idx in range(2, ws_rank.max_row + 1):
            rating_cell = ws_rank.cell(row=row_idx, column=10)  # Column J = Rating
            fill = _RATING_COLORS.get(str(rating_cell.value))
            if fill:
                rating_cell.fill = fill
                if fill in (_RED,):
                    rating_cell.font = _WHITE_FONT

        # Color-code pattern alerts (column A = Alert Type)
        if "Pattern Alerts" in writer.book.sheetnames:
            ws_alerts = writer.book["Pattern Alerts"]
            for row_idx in range(2, ws_alerts.max_row + 1):
                type_cell = ws_alerts.cell(row=row_idx, column=1)
                fill = _ALERT_COLORS.get(str(type_cell.value))
                if fill:
                    type_cell.fill = fill
                    if fill in (_RED,):
                        type_cell.font = _WHITE_FONT

    return filepath


def _write_target_return_excel(target_return_output, out_path: Path) -> Path:
    """Write Agent 14 Target Return Constructor output to Excel."""
    today = date.today().isoformat()
    filepath = out_path / f"agent14_target_return_{today}.xlsx"

    # --- Summary ---
    ta = target_return_output.tier_allocation
    pa = target_return_output.probability_assessment
    rd = target_return_output.risk_disclosure
    summary_rows = [
        {"Field": "Portfolio Name", "Value": target_return_output.portfolio_name},
        {"Field": "Target Return %", "Value": f"{target_return_output.target_return_pct:.1f}%"},
        {"Field": "Time Horizon", "Value": f"{target_return_output.time_horizon_months} months"},
        {"Field": "Total Capital", "Value": f"${target_return_output.total_capital:,.0f}"},
        {"Field": "Market Regime", "Value": target_return_output.market_regime},
        {"Field": "Achievability", "Value": rd.achievability_rating},
        {"Field": "Prob Achieve Target", "Value": f"{pa.prob_achieve_target:.1%}"},
        {"Field": "Expected Return", "Value": f"{pa.expected_return_pct:.1f}%"},
        {"Field": "Median Return", "Value": f"{pa.median_return_pct:.1f}%"},
        {"Field": "Cash Reserve", "Value": f"{target_return_output.cash_reserve_pct:.1f}%"},
        {"Field": "Position Count", "Value": len(target_return_output.positions)},
        {"Field": "Analysis Date", "Value": target_return_output.analysis_date},
        {"Field": "", "Value": ""},
        {"Field": "Summary", "Value": target_return_output.summary},
    ]
    df_summary = pd.DataFrame(summary_rows)

    # --- Positions ---
    pos_rows = []
    for p in target_return_output.positions:
        pos_rows.append({
            "Ticker": p.ticker,
            "Company": p.company_name,
            "Tier": p.tier,
            "Allocation %": f"{p.allocation_pct:.2f}%",
            "Dollar Amount": f"${p.dollar_amount:,.0f}",
            "Shares": p.shares,
            "Entry Strategy": p.entry_strategy,
            "Entry Price": f"${p.target_entry_price:.2f}",
            "Stop Loss": f"${p.stop_loss_price:.2f}",
            "Stop Loss %": f"{p.stop_loss_pct:.1f}%",
            "Return Contrib %": f"{p.expected_return_contribution_pct:.2f}%",
            "Conviction": p.conviction_level,
            "Composite": p.composite_score,
            "EPS": p.eps_rating,
            "RS": p.rs_rating,
            "Sector": p.sector,
            "Rationale": p.selection_rationale,
        })
    df_positions = pd.DataFrame(pos_rows)

    # --- Tier Mix ---
    tier_rows = [
        {"Tier": "T1 Momentum", "Allocation %": f"{ta.t1_momentum_pct:.1f}%"},
        {"Tier": "T2 Quality Growth", "Allocation %": f"{ta.t2_quality_growth_pct:.1f}%"},
        {"Tier": "T3 Defensive", "Allocation %": f"{ta.t3_defensive_pct:.1f}%"},
        {"Tier": "Cash", "Allocation %": f"{target_return_output.cash_reserve_pct:.1f}%"},
        {"Tier": "", "Allocation %": ""},
        {"Tier": "Rationale", "Allocation %": ta.rationale},
    ]
    df_tiers = pd.DataFrame(tier_rows)

    # --- Scenarios ---
    scenarios = target_return_output.scenarios
    scen_rows = []
    for s in [scenarios.bull_scenario, scenarios.base_scenario, scenarios.bear_scenario]:
        scen_rows.append({
            "Scenario": s.name,
            "Probability %": f"{s.probability_pct:.1f}%",
            "Portfolio Return %": f"{s.portfolio_return_pct:.1f}%",
            "S&P 500 %": f"{s.sp500_return_pct:.1f}%",
            "NASDAQ %": f"{s.nasdaq_return_pct:.1f}%",
            "DOW %": f"{s.dow_return_pct:.1f}%",
            "Alpha vs S&P": f"{s.alpha_vs_sp500:.1f}%",
            "Max Drawdown %": f"{s.max_drawdown_pct:.1f}%",
            "Stops Triggered": s.stops_triggered,
            "Description": s.description,
        })
    df_scenarios = pd.DataFrame(scen_rows)

    # --- Probability Assessment ---
    prob_rows = [
        {"Metric": "Prob Achieve Target", "Value": f"{pa.prob_achieve_target:.1%}"},
        {"Metric": "Prob Positive Return", "Value": f"{pa.prob_positive_return:.1%}"},
        {"Metric": "Prob Beat S&P 500", "Value": f"{pa.prob_beat_sp500:.1%}"},
        {"Metric": "Prob Beat NASDAQ", "Value": f"{pa.prob_beat_nasdaq:.1%}"},
        {"Metric": "Prob Beat DOW", "Value": f"{pa.prob_beat_dow:.1%}"},
        {"Metric": "", "Value": ""},
        {"Metric": "Expected Return", "Value": f"{pa.expected_return_pct:.1f}%"},
        {"Metric": "Median Return", "Value": f"{pa.median_return_pct:.1f}%"},
        {"Metric": "P10 Return", "Value": f"{pa.p10_return_pct:.1f}%"},
        {"Metric": "P25 Return", "Value": f"{pa.p25_return_pct:.1f}%"},
        {"Metric": "P50 Return", "Value": f"{pa.p50_return_pct:.1f}%"},
        {"Metric": "P75 Return", "Value": f"{pa.p75_return_pct:.1f}%"},
        {"Metric": "P90 Return", "Value": f"{pa.p90_return_pct:.1f}%"},
    ]
    df_probability = pd.DataFrame(prob_rows)

    # --- Alternatives ---
    alt_rows = []
    for a in target_return_output.alternatives:
        alt_rows.append({
            "Name": a.name,
            "Target Return %": f"{a.target_return_pct:.1f}%",
            "Prob Achieve": f"{a.prob_achieve_target:.1%}",
            "Positions": a.position_count,
            "T1 %": f"{a.t1_pct:.0f}%",
            "T2 %": f"{a.t2_pct:.0f}%",
            "T3 %": f"{a.t3_pct:.0f}%",
            "Max Drawdown %": f"{a.max_drawdown_pct:.1f}%",
            "Key Difference": a.key_difference,
            "Tradeoff": a.tradeoff,
        })
    df_alternatives = pd.DataFrame(alt_rows)

    # --- Risk Disclosure ---
    risk_rows = [
        {"Field": "Achievability Rating", "Value": rd.achievability_rating},
        {"Field": "Achievability Rationale", "Value": rd.achievability_rationale},
        {"Field": "Max Expected Drawdown", "Value": f"{rd.max_expected_drawdown_pct:.1f}%"},
        {"Field": "Recovery Time (months)", "Value": f"{rd.recovery_time_months:.1f}"},
        {"Field": "", "Value": ""},
        {"Field": "Conditions for Success", "Value": " | ".join(rd.conditions_for_success)},
        {"Field": "Conditions for Failure", "Value": " | ".join(rd.conditions_for_failure)},
        {"Field": "", "Value": ""},
        {"Field": "Disclaimer", "Value": rd.disclaimer},
        {"Field": "", "Value": ""},
        {"Field": "Construction Rationale", "Value": target_return_output.construction_rationale},
    ]
    df_risk = pd.DataFrame(risk_rows)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        df_positions.to_excel(writer, sheet_name="Positions", index=False)
        df_tiers.to_excel(writer, sheet_name="Tier Mix", index=False)
        df_scenarios.to_excel(writer, sheet_name="Scenarios", index=False)
        df_probability.to_excel(writer, sheet_name="Probability", index=False)
        df_alternatives.to_excel(writer, sheet_name="Alternatives", index=False)
        df_risk.to_excel(writer, sheet_name="Risk Disclosure", index=False)

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    return filepath


# ---------------------------------------------------------------------------
# Token Usage Excel Writer
# ---------------------------------------------------------------------------

def _write_token_usage_excel(output_dir: Path) -> str | None:
    """Write token usage report to Excel if any LLM calls were tracked."""
    from ibd_agents.tools.token_tracker import _compute_cost

    if not token_tracker.has_records:
        return None

    filepath = str(output_dir / "token_usage.xlsx")
    summary = token_tracker.get_summary()
    by_agent = token_tracker.get_by_agent()
    by_function = token_tracker.get_by_function()

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        # --- Sheet 1: Summary ---
        summary_data = [
            {"Metric": "Total Input Tokens", "Value": f"{summary['total_input_tokens']:,}"},
            {"Metric": "Total Output Tokens", "Value": f"{summary['total_output_tokens']:,}"},
            {"Metric": "Total Tokens", "Value": f"{summary['total_tokens']:,}"},
            {"Metric": "Estimated Cost ($)", "Value": f"${summary['estimated_cost_usd']:.4f}"},
            {"Metric": "Number of LLM Calls", "Value": str(summary["num_calls"])},
            {"Metric": "Model", "Value": "claude-haiku-3.5"},
            {"Metric": "Pipeline Run Date", "Value": str(date.today())},
        ]
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        ws = writer.sheets["Summary"]
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 25

        # --- Sheet 2: By Agent ---
        agent_rows = []
        for a in by_agent:
            agent_rows.append({
                "Agent": a["agent_id"],
                "Name": a["agent_name"],
                "Input Tokens": a["input_tokens"],
                "Output Tokens": a["output_tokens"],
                "Total Tokens": a["total_tokens"],
                "Cost ($)": a["cost_usd"],
                "Calls": a["calls"],
            })
        # Totals row
        agent_rows.append({
            "Agent": "TOTAL",
            "Name": "",
            "Input Tokens": summary["total_input_tokens"],
            "Output Tokens": summary["total_output_tokens"],
            "Total Tokens": summary["total_tokens"],
            "Cost ($)": summary["estimated_cost_usd"],
            "Calls": summary["num_calls"],
        })
        df_agent = pd.DataFrame(agent_rows)
        df_agent.to_excel(writer, sheet_name="By Agent", index=False)
        ws = writer.sheets["By Agent"]
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        # Bold totals row
        total_row = len(agent_rows) + 1
        for cell in ws[total_row]:
            cell.font = Font(bold=True)
        # Format number columns
        from openpyxl.utils import get_column_letter
        for col_idx in [3, 4, 5, 7]:  # Input, Output, Total, Calls
            for row in range(2, total_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.number_format = "#,##0"
        for row in range(2, total_row + 1):
            ws.cell(row=row, column=6).number_format = "$#,##0.0000"
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 25)

        # --- Sheet 3: By Function ---
        func_rows = []
        for f in by_function:
            func_rows.append({
                "Function": f["function"],
                "Agent": f["agent_id"],
                "Name": f["agent_name"],
                "Input Tokens": f["input_tokens"],
                "Output Tokens": f["output_tokens"],
                "Total Tokens": f["total_tokens"],
                "Cost ($)": f["cost_usd"],
                "Calls": f["calls"],
            })
        # Totals row
        func_rows.append({
            "Function": "TOTAL",
            "Agent": "",
            "Name": "",
            "Input Tokens": summary["total_input_tokens"],
            "Output Tokens": summary["total_output_tokens"],
            "Total Tokens": summary["total_tokens"],
            "Cost ($)": summary["estimated_cost_usd"],
            "Calls": summary["num_calls"],
        })
        df_func = pd.DataFrame(func_rows)
        df_func.to_excel(writer, sheet_name="By Function", index=False)
        ws = writer.sheets["By Function"]
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        total_row = len(func_rows) + 1
        for cell in ws[total_row]:
            cell.font = Font(bold=True)
        for col_idx in [4, 5, 6, 8]:  # Input, Output, Total, Calls
            for row in range(2, total_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.number_format = "#,##0"
        for row in range(2, total_row + 1):
            ws.cell(row=row, column=7).number_format = "$#,##0.0000"
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    return filepath


def main(data_dir: str = "data", output_dir: str = "output",
         start_from: int = 1, run_historical: bool = False):
    out_path = Path(output_dir)
    out_path.mkdir(exist_ok=True)
    snapshots_dir = out_path / "snapshots"
    snapshots_dir.mkdir(exist_ok=True)

    def should_run(agent_num: int) -> bool:
        # Agents 11/12 now run before Agent 05 in the pipeline.
        # When resuming from Agent 05+, load 11/12 from snapshots.
        if agent_num in (11, 12) and start_from > 4 and start_from not in (11, 12):
            return False
        # Agent 14 runs after Agent 07. When resuming from Agent 08+,
        # load 14 from snapshot (unless specifically targeting 14).
        if agent_num == 14 and start_from > 7 and start_from != 14:
            return False
        # Agent 16 runs after Agent 02. When resuming from Agent 03+,
        # load 16 from snapshot (unless specifically targeting 16).
        if agent_num == 16 and start_from > 2 and start_from != 16:
            return False
        # Agent 17 runs after Agent 15. When resuming from Agent 08+,
        # load 17 from snapshot (unless specifically targeting 17).
        if agent_num == 17 and start_from > 7 and start_from != 17:
            return False
        return agent_num >= start_from

    if start_from > 1:
        print(f"\n=== Resuming from Agent {start_from:02d} — "
              f"loading agents 1..{start_from - 1} from snapshots ===\n")

    # ===== PHASE 1: Research Agent =====
    if should_run(1):
        print(f"[Agent 01] Running Research pipeline against '{data_dir}' ...")
        research_output = run_research_pipeline(data_dir)
        print(f"[Agent 01] Done — {len(research_output.stocks)} stocks, {len(research_output.etfs)} ETFs")
        _save_snapshot(research_output, 1, snapshots_dir)
        research_file = _write_research_excel(research_output, out_path)
        print(f"[Agent 01] Saved: {research_file}")
    else:
        research_output = _load_snapshot(1, snapshots_dir)

    # ===== PHASE 2: Analyst Agent =====
    if should_run(2):
        print(f"[Agent 02] Running Analyst pipeline ...")
        analyst_output = run_analyst_pipeline(research_output)
        print(f"[Agent 02] Done — {len(analyst_output.rated_stocks)} top stocks, "
              f"{len(analyst_output.ibd_keeps)} IBD keeps, "
              f"{len(analyst_output.sector_rankings)} sectors ranked")
        _save_snapshot(analyst_output, 2, snapshots_dir)
        analyst_file = _write_analyst_excel(analyst_output, out_path, research_output)
        print(f"[Agent 02] Saved: {analyst_file}")
    else:
        analyst_output = _load_snapshot(2, snapshots_dir)

    # ===== PHASE 2b: Regime Detector =====
    regime_output = None
    if should_run(16):
        print(f"[Agent 16] Running Regime Detector pipeline ...")
        regime_output = run_regime_detector_pipeline(
            analyst_output=analyst_output,
            use_real_data=False,
        )
        legacy = regime_output.to_legacy_regime()
        print(f"[Agent 16] Done — regime={regime_output.regime.value}, "
              f"confidence={regime_output.confidence.value}, "
              f"health={regime_output.market_health_score}/10, "
              f"exposure={regime_output.exposure_recommendation}%, "
              f"legacy={legacy}")
        _save_snapshot(regime_output, 16, snapshots_dir)
        regime_file = _write_regime_detector_excel(regime_output, out_path)
        print(f"[Agent 16] Saved: {regime_file}")
    elif start_from > 2:
        try:
            regime_output = _load_snapshot(16, snapshots_dir)
        except SystemExit:
            print(f"  [snapshot] Agent 16 snapshot not found — skipping regime detector output")
            regime_output = None

    # ===== PHASE 2a: Historical Analyst (opt-in, runs early for momentum data) =====
    historical_output = None
    if run_historical:
        if should_run(13):
            print(f"[Agent 13] Running Historical Analyst pipeline ...")
            try:
                from ibd_agents.tools.historical_store import ingest_directory
                for ingest_dir in ["data/ibd_pdf", "data/ibd_xls", "data/ibd_history"]:
                    if Path(ingest_dir).exists():
                        ingest_result = ingest_directory(ingest_dir)
                        if ingest_result.get("total_stocks_added", 0) > 0:
                            print(f"[Agent 13] Ingested {ingest_result['total_stocks_added']} "
                                  f"stocks from {ingest_dir}/")

                historical_output = run_historical_pipeline(analyst_output)
                _save_snapshot(historical_output, 13, snapshots_dir)
                print(f"[Agent 13] Done — source={historical_output.historical_source}, "
                      f"{len(historical_output.stock_analyses)} stocks analyzed, "
                      f"{len(historical_output.improving_stocks)} improving, "
                      f"{len(historical_output.deteriorating_stocks)} deteriorating, "
                      f"{len(historical_output.historical_analogs)} analogs")
            except Exception as e:
                print(f"[Agent 13] Warning: Historical analysis failed: {e}")
                print(f"[Agent 13] Install chromadb to enable historical context analysis")
        else:
            historical_output = _load_snapshot(13, snapshots_dir)
    else:
        print(f"[Agent 13] Skipped (use --historical to enable)")

    # ===== PHASE 3: Rotation Detector =====
    if should_run(3):
        print(f"[Agent 03] Running Rotation Detection pipeline ...")
        rotation_output = run_rotation_pipeline(analyst_output, research_output)
        print(f"[Agent 03] Done — verdict={rotation_output.verdict.value}, "
              f"{rotation_output.signals.signals_active}/5 signals, "
              f"type={rotation_output.rotation_type.value}")
        _save_snapshot(rotation_output, 3, snapshots_dir)
        rotation_file = _write_rotation_excel(rotation_output, out_path)
        print(f"[Agent 03] Saved: {rotation_file}")
    else:
        rotation_output = _load_snapshot(3, snapshots_dir)

    # ===== PHASE 4: Sector Strategist =====
    if should_run(4):
        print(f"[Agent 04] Running Sector Strategy pipeline ...")
        strategy_output = run_strategist_pipeline(rotation_output, analyst_output)
        print(f"[Agent 04] Done — "
              f"{len(strategy_output.sector_allocations.overall_allocation)} sectors allocated, "
              f"{len(strategy_output.theme_recommendations)} themes, "
              f"{len(strategy_output.rotation_signals)} signals")
        _save_snapshot(strategy_output, 4, snapshots_dir)
        strategy_file = _write_strategy_excel(strategy_output, out_path)
        print(f"[Agent 04] Saved: {strategy_file}")
    else:
        strategy_output = _load_snapshot(4, snapshots_dir)

    # ===== PHASE 4a: Value Investor (runs before portfolio for integration) =====
    if should_run(11):
        print(f"[Agent 11] Running Value Investor pipeline ...")
        value_output = run_value_investor_pipeline(analyst_output)

        cat_counts = {}
        for vs in value_output.value_stocks:
            cat_counts[vs.value_category] = cat_counts.get(vs.value_category, 0) + 1
        cat_str = ", ".join(f"{c}: {n}" for c, n in sorted(cat_counts.items(), key=lambda x: -x[1]))
        print(f"[Agent 11] Done — {len(value_output.value_stocks)} stocks scored, "
              f"{len(value_output.value_traps)} value traps, "
              f"{len(value_output.top_value_picks)} top picks")
        print(f"[Agent 11] Categories: {cat_str}")
        _save_snapshot(value_output, 11, snapshots_dir)

        value_file = _write_value_investor_excel(value_output, out_path)
        print(f"[Agent 11] Saved: {value_file}")
    else:
        value_output = _load_snapshot(11, snapshots_dir)

    # ===== PHASE 4b: PatternAlpha Per-Stock Scoring (runs before portfolio) =====
    if should_run(12):
        print(f"[Agent 12] Running PatternAlpha per-stock scoring pipeline ...")
        pattern_output = run_pattern_pipeline(analyst_output)
        ps = pattern_output.portfolio_summary
        print(f"[Agent 12] Done — {ps.total_stocks_scored} stocks scored, "
              f"{ps.stocks_with_patterns} with patterns, "
              f"avg enhanced={ps.avg_enhanced_score:.1f}/150, "
              f"{len(pattern_output.pattern_alerts)} alerts, "
              f"source={pattern_output.scoring_source}")
        _save_snapshot(pattern_output, 12, snapshots_dir)

        pattern_file = _write_pattern_excel(pattern_output, out_path)
        print(f"[Agent 12] Saved: {pattern_file}")
    else:
        pattern_output = _load_snapshot(12, snapshots_dir)

    # ===== PHASE 5: Portfolio Manager =====
    if should_run(5):
        print(f"[Agent 05] Running Portfolio Manager pipeline ...")
        portfolio_output = run_portfolio_pipeline(
            strategy_output, analyst_output,
            value_output=value_output, pattern_output=pattern_output,
            historical_output=historical_output,
        )
        print(f"[Agent 05] Done — {portfolio_output.total_positions} positions, "
              f"{portfolio_output.stock_count} stocks, {portfolio_output.etf_count} ETFs, "
              f"{portfolio_output.keeps_placement.total_keeps} keeps")
        _save_snapshot(portfolio_output, 5, snapshots_dir)
        portfolio_file = _write_portfolio_excel(portfolio_output, out_path)
        print(f"[Agent 05] Saved: {portfolio_file}")
    else:
        portfolio_output = _load_snapshot(5, snapshots_dir)

    # ===== PHASE 6: Risk Officer =====
    if should_run(6):
        print(f"[Agent 06] Running Risk Officer pipeline ...")
        risk_output = run_risk_pipeline(portfolio_output, strategy_output, analyst_output=analyst_output, pattern_output=pattern_output)
        print(f"[Agent 06] Done — {risk_output.overall_status}, "
              f"Sleep Well: {risk_output.sleep_well_scores.overall_score}/10, "
              f"{len(risk_output.vetoes)} vetoes, {len(risk_output.warnings)} warnings")
        _save_snapshot(risk_output, 6, snapshots_dir)
        risk_file = _write_risk_excel(risk_output, out_path)
        print(f"[Agent 06] Saved: {risk_file}")
    else:
        risk_output = _load_snapshot(6, snapshots_dir)

    # ===== PHASE 15: Exit Strategist =====
    exit_strategy_output = None
    if should_run(15):
        # Read actual brokerage holdings for exit signal evaluation
        brokerage_holdings = None
        try:
            brokerage_holdings = read_brokerage_pdfs("data/portfolios")
            print(f"[Agent 15] Loaded {len(brokerage_holdings.holdings)} brokerage positions "
                  f"from {brokerage_holdings.account_count} accounts")
        except Exception as e:
            print(f"[Agent 15] No brokerage PDFs loaded ({e}) — using model portfolio only")

        print(f"[Agent 15] Running Exit Strategist pipeline ...")
        exit_strategy_output = run_exit_strategist_pipeline(
            portfolio_output, risk_output, strategy_output, analyst_output, rotation_output,
            brokerage_holdings=brokerage_holdings,
        )
        n_crit = sum(1 for s in exit_strategy_output.signals if s.urgency.value == "CRITICAL")
        n_warn = sum(1 for s in exit_strategy_output.signals if s.urgency.value == "WARNING")
        n_healthy = sum(1 for s in exit_strategy_output.signals if s.urgency.value == "HEALTHY")
        print(f"[Agent 15] Done — {len(exit_strategy_output.signals)} signals, "
              f"health={exit_strategy_output.portfolio_health_score}/10, "
              f"{n_crit} critical, {n_warn} warning, {n_healthy} healthy")
        _save_snapshot(exit_strategy_output, 15, snapshots_dir)
        exit_file = _write_exit_strategy_excel(exit_strategy_output, out_path)
        print(f"[Agent 15] Saved: {exit_file}")
    elif start_from > 6:
        try:
            exit_strategy_output = _load_snapshot(15, snapshots_dir)
        except SystemExit:
            print(f"  [snapshot] Agent 15 snapshot not found — skipping exit strategy output")
            exit_strategy_output = None

    # ===== PHASE 17: Earnings Risk Analyst =====
    earnings_risk_output = None
    if should_run(17):
        # Re-use brokerage_holdings from Agent 15 (or load fresh)
        try:
            brokerage_holdings  # noqa: F841 — check if defined
        except NameError:
            brokerage_holdings = None
        if brokerage_holdings is None:
            try:
                brokerage_holdings = read_brokerage_pdfs("data/portfolios")
                print(f"[Agent 17] Loaded {len(brokerage_holdings.holdings)} brokerage positions "
                      f"from {brokerage_holdings.account_count} accounts")
            except Exception as e:
                print(f"[Agent 17] No brokerage PDFs loaded ({e}) — using model portfolio only")

        print(f"[Agent 17] Running Earnings Risk Analyst pipeline ...")
        earnings_risk_output = run_earnings_risk_pipeline(
            portfolio_output,
            regime_output=regime_output,
            analyst_output=analyst_output,
            brokerage_holdings=brokerage_holdings,
        )
        n_crit = sum(1 for a in earnings_risk_output.analyses if a.risk_level.value == "CRITICAL")
        n_high = sum(1 for a in earnings_risk_output.analyses if a.risk_level.value == "HIGH")
        n_mod = sum(1 for a in earnings_risk_output.analyses if a.risk_level.value == "MODERATE")
        n_low = sum(1 for a in earnings_risk_output.analyses if a.risk_level.value == "LOW")
        print(f"[Agent 17] Done — {len(earnings_risk_output.analyses)} positions with earnings, "
              f"{len(earnings_risk_output.positions_clear)} clear, "
              f"{n_crit} critical, {n_high} high, {n_mod} moderate, {n_low} low, "
              f"concentration={earnings_risk_output.concentration.concentration_risk}")
        _save_snapshot(earnings_risk_output, 17, snapshots_dir)
        earnings_risk_file = _write_earnings_risk_excel(earnings_risk_output, out_path)
        print(f"[Agent 17] Saved: {earnings_risk_file}")
    elif start_from > 7:
        try:
            earnings_risk_output = _load_snapshot(17, snapshots_dir)
        except SystemExit:
            print(f"  [snapshot] Agent 17 snapshot not found — skipping earnings risk output")
            earnings_risk_output = None

    # ===== PHASE 7: Returns Projector =====
    if should_run(7):
        print(f"[Agent 07] Running Returns Projector pipeline ...")
        returns_output = run_returns_projector_pipeline(
            portfolio_output, rotation_output, risk_output, analyst_output,
        )
        print(f"[Agent 07] Done — Expected 12m: {returns_output.expected_return.expected_12m:.1f}%, "
              f"Alpha vs SPY: {returns_output.alpha_analysis.net_expected_alpha_vs_spy:.1f}%, "
              f"DD w/stops: {returns_output.risk_metrics.max_drawdown_with_stops:.1f}%")
        _save_snapshot(returns_output, 7, snapshots_dir)
        returns_file = _write_returns_projector_excel(returns_output, out_path)
        print(f"[Agent 07] Saved: {returns_file}")
    else:
        returns_output = _load_snapshot(7, snapshots_dir)

    # ===== PHASE 7a: Target Return Constructor =====
    target_return_output = None
    if should_run(14):
        print(f"[Agent 14] Running Target Return Constructor pipeline ...")
        target_return_output = run_target_return_pipeline(
            analyst_output, rotation_output, strategy_output,
            risk_output, returns_output,
        )
        print(f"[Agent 14] Done — {len(target_return_output.positions)} positions, "
              f"prob={target_return_output.probability_assessment.prob_achieve_target:.0%}, "
              f"achievability={target_return_output.risk_disclosure.achievability_rating}")
        _save_snapshot(target_return_output, 14, snapshots_dir)
        target_return_file = _write_target_return_excel(target_return_output, out_path)
        print(f"[Agent 14] Saved: {target_return_file}")
    elif start_from > 7:
        try:
            target_return_output = _load_snapshot(14, snapshots_dir)
        except SystemExit:
            print(f"  [snapshot] Agent 14 snapshot not found — skipping target return output")
            target_return_output = None

    # ===== PHASE 8: Portfolio Reconciler =====
    if should_run(8):
        print(f"[Agent 08] Running Reconciler pipeline ...")
        reconciler_output = run_reconciler_pipeline(
            portfolio_output, analyst_output, returns_output=returns_output,
            rotation_output=rotation_output, strategy_output=strategy_output,
            risk_output=risk_output,
            portfolios_dir="data/portfolios",
            value_output=value_output, pattern_output=pattern_output,
        )
        print(f"[Agent 08] Done — {len(reconciler_output.actions)} actions, "
              f"turnover: {reconciler_output.transformation_metrics.turnover_pct:.0f}%")
        _save_snapshot(reconciler_output, 8, snapshots_dir)
        reconciler_file = _write_reconciler_excel(reconciler_output, out_path)
        print(f"[Agent 08] Saved: {reconciler_file}")
    else:
        reconciler_output = _load_snapshot(8, snapshots_dir)

    # ===== PHASE 9: Educator =====
    if should_run(9):
        print(f"[Agent 09] Running Educator pipeline ...")
        educator_output = run_educator_pipeline(
            portfolio_output, analyst_output, rotation_output, strategy_output,
            risk_output, reconciler_output, returns_output=returns_output,
        )
        print(f"[Agent 09] Done — {len(educator_output.stock_explanations)} stocks explained, "
              f"{len(educator_output.concept_lessons)} lessons, "
              f"{len(educator_output.glossary)} glossary entries")
        _save_snapshot(educator_output, 9, snapshots_dir)
        educator_file = _write_educator_excel(educator_output, out_path)
        print(f"[Agent 09] Saved: {educator_file}")
    else:
        educator_output = _load_snapshot(9, snapshots_dir)

    # ===== PHASE 10: Executive Summary Synthesizer =====
    if should_run(10):
        print(f"[Agent 10] Running Executive Summary Synthesizer pipeline ...")
        synthesis_output = run_synthesizer_pipeline(
            research_output, analyst_output, rotation_output, strategy_output,
            portfolio_output, risk_output, returns_output, reconciler_output,
            educator_output,
        )
        print(f"[Agent 10] Done — source={synthesis_output.synthesis_source}, "
              f"{len(synthesis_output.cross_agent_connections)} connections, "
              f"{len(synthesis_output.contradictions)} contradictions")
        _save_snapshot(synthesis_output, 10, snapshots_dir)
        synthesis_file = _write_synthesizer_excel(synthesis_output, out_path)
        print(f"[Agent 10] Saved: {synthesis_file}")
    else:
        synthesis_output = _load_snapshot(10, snapshots_dir)

    # ===== PDF Summary Report =====
    print(f"\n[PDF] Generating summary report ...")
    try:
        from ibd_agents.reports.pdf_generator import generate_pdf_report
        pdf_file = generate_pdf_report(
            research_output, analyst_output, rotation_output, strategy_output,
            portfolio_output, risk_output, returns_output, reconciler_output,
            educator_output, synthesis_output, value_output, pattern_output,
            out_path, historical_output=historical_output,
            target_return_output=target_return_output,
        )
        print(f"[PDF] Saved: {pdf_file}")
    except Exception as e:
        print(f"[PDF] Warning: PDF generation failed: {e}")
        print(f"[PDF] Excel files are still available in {out_path}/")

    # ===== Token Usage Report =====
    token_file = _write_token_usage_excel(out_path)
    if token_file:
        summary = token_tracker.get_summary()
        print(f"\n[Tokens] {summary['num_calls']} LLM calls, "
              f"{summary['total_tokens']:,} tokens, "
              f"${summary['estimated_cost_usd']:.4f}")
        print(f"[Tokens] Saved: {token_file}")

    print(f"\nOutput directory: {out_path}/")


if __name__ == "__main__":
    args = parse_args()
    main(
        data_dir=args.data,
        output_dir=args.output,
        start_from=args.start_from,
        run_historical=args.historical,
    )
